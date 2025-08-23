from django.shortcuts import render, redirect, get_object_or_404
from rest_framework import generics
from django.db import models
from .models import Teacher, Homework, Gallery, ContactMessage, Result, Notice, Student, HomeworkSubmission, ActivityLog, StudentAccount, LeadershipMessage, Subject, Marks, Resource
from .serializers import TeacherSerializer, HomeworkSerializer, GallerySerializer, ContactMessageSerializer, ResultSerializer, NoticeSerializer, StudentSerializer, RegisterSerializer, HomeworkSubmissionSerializer, ActivityLogSerializer, GalleryLikeSerializer, GalleryCommentSerializer, SubjectSerializer, MarksSerializer, StudentResultSerializer, ResourceSerializer

# Import website content views
from . import views_website_content
from rest_framework import permissions
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.contrib.auth import authenticate, login, logout
from rest_framework.permissions import IsAuthenticated
from django.utils import timezone
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden
from django import forms
from .models import ClassSection
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.contrib.auth.models import User
from django.views.decorators.http import require_GET
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
# PDF and Excel export imports
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import xlsxwriter
from io import BytesIO
import os
from datetime import datetime, date
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

import io
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import json
from datetime import datetime

# Email functionality imports
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.conf import settings
import threading
from django.core.mail import send_mail
from django.template import Context, Template

# Import website content views
from .views_website_content import index_page, about_page, contact_page

def faculty_page(request):
    return render(request, 'faculty.html')

def gallery_page(request):
    from .models import Gallery, GalleryLike, GalleryComment
    from django.shortcuts import render
    page = int(request.GET.get('page', 1))
    per_page = 9
    gallery_qs = Gallery.objects.filter(status='Published').order_by('-upload_date')
    paginator = Paginator(gallery_qs, per_page)
    gallery_items = paginator.get_page(page)
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        # AJAX infinite scroll
        data = []
        for item in gallery_items:
            item_data = {
                'id': item.id,
                'caption': item.caption,
                'description': item.description,
                'category': item.category,
                'upload_date': item.upload_date.strftime('%Y-%m-%d'),
                'like_count': item.likes.count(),
                'comment_count': item.comments.count(),
                'media_type': getattr(item, 'media_type', 'image'),
            }
            
            # Add media URL based on media type
            if hasattr(item, 'media_url'):
                item_data['media_url'] = item.media_url
            else:
                # Fallback for existing items without media_url property
                if item.media_type == 'video' and item.video:
                    item_data['image_url'] = item.video.url
                elif item.media_type == 'image' and item.image:
                    item_data['image_url'] = item.image.url
                else:
                    item_data['image_url'] = ''
                
            data.append(item_data)
        return JsonResponse({'gallery': data, 'has_next': gallery_items.has_next()})
    # For initial page load
    for item in gallery_items:
        item.like_count = item.likes.count()
        item.comment_count = item.comments.count()
    
    return render(request, 'gallery.html', {'gallery_items': gallery_items})

@login_required
def gallery_upload(request):
    from .models import Gallery
    from django.contrib import messages
    from django.shortcuts import redirect, render
    
    if request.method == 'POST':
        media_type = request.POST.get('media_type')
        caption = request.POST.get('caption')
        description = request.POST.get('description', '')
        category = request.POST.get('category', '')
        tags = request.POST.get('tags', '')
        status = request.POST.get('status', 'Draft')
        
        # Create new gallery item
        gallery_item = Gallery(
            media_type=media_type,
            caption=caption,
            description=description,
            category=category,
            tags=tags,
            status=status,
            uploader=request.user
        )
        
        # Handle file upload based on media type
        if media_type == 'image' and 'image' in request.FILES:
            gallery_item.image = request.FILES['image']
        elif media_type == 'video' and 'video' in request.FILES:
            gallery_item.video = request.FILES['video']
        else:
            messages.error(request, 'Please upload a valid file.')
            return render(request, 'gallery_upload.html')
        
        # Save the gallery item
        gallery_item.save()
        messages.success(request, 'Your media has been uploaded successfully!')
        return redirect('gallery')
    
    return render(request, 'gallery_upload.html')
    # Determine display name from Student or Teacher profile if available
    if request.user.is_authenticated:
        profile_name = None
        try:
            from .models import Student, Teacher
            student = Student.objects.filter(user=request.user).first()
            if student:
                profile_name = student.full_name
            else:
                teacher = Teacher.objects.filter(profile__user=request.user).first()
                if teacher:
                    profile_name = teacher.full_name
        except Exception:
            profile_name = None
        if not profile_name:
            profile_name = request.user.get_full_name() or request.user.username
    else:
        profile_name = 'Anonymous'
    return render(request, 'gallery.html', {'gallery_items': gallery_items, 'user_name': profile_name})

# About and Contact page views are now imported from views_website_content

def homework_page(request):
    return render(request, 'homework.html')

def library_page(request):
    return render(request, 'library.html')

def notices_page(request):
    return render(request, 'notices.html')

def notes_page(request):
    """Public page that forwards students to dashboard resources section or provides a simple list."""
    return render(request, 'notes.html')

def results_page(request):
    return render(request, 'results.html')

def student_signup_page(request):
    return render(request, 'student_signup.html')

def welcome_page(request):
    return render(request, 'welcome.html')

def login_page(request):
    next_url = request.GET.get('next', '/')
    msg = ''
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect(request.POST.get('next', '/'))
        else:
            msg = "Invalid username or password."
    return render(request, 'login.html', {'msg': msg, 'next': next_url})

def student_dashboard(request):
    from .models import StatusUpdate, Post
    user = request.user
    # Handle status update form
    if request.method == 'POST' and 'status_text' in request.POST:
        text = request.POST.get('status_text', '').strip()
        image = request.FILES.get('status_image')
        if text or image:
            StatusUpdate.objects.create(user=user, text=text, image=image)
        return redirect('student_dashboard')
    # Fetch feed: status updates and posts by user and followed users
    feed = list(StatusUpdate.objects.filter(user=user))
    # Optionally, add posts from followed users (if Follow model is used)
    # For now, just show user's own updates
    feed += list(Post.objects.filter(user=user))
    # Sort by created_at descending
    feed = sorted(feed, key=lambda x: x.created_at, reverse=True)
    return render(request, 'student_dashboard.html', {'feed': feed, 'user': user})

def teacher_dashboard_view(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    # Check if user is a teacher
    if not hasattr(request.user, 'teacher_profile'):
        return redirect('index')
    
    from .models import StatusUpdate, Homework, Student, Attendance, ClassSection
    from datetime import date
    
    teacher = request.user.teacher_profile
    
    # Get statistics
    assigned_classes = ClassSection.objects.filter(teachers=teacher)
    total_students = Student.objects.filter(class_section__in=assigned_classes).count()
    total_homeworks = Homework.objects.filter(assigned_class__in=assigned_classes.values_list('name', flat=True)).count()
    total_posts = StatusUpdate.objects.filter(user=request.user).count()
    
    # Get today's attendance
    today = date.today()
    attendance_today = Attendance.objects.filter(
        class_section__in=assigned_classes,
        date=today,
        status='present'
    ).count()
    
    # Get all posts for the feed
    posts = StatusUpdate.objects.all().order_by('-created_at')
    
    # Get teacher's homeworks
    homeworks = Homework.objects.filter(assigned_class__in=assigned_classes.values_list('name', flat=True)).order_by('-id')[:10]
    
    # Get assigned students
    students = Student.objects.filter(class_section__in=assigned_classes).order_by('full_name')
    
    # Handle post creation (from the feed form)
    if request.method == 'POST' and request.POST.get('form_type') == 'post':
        print(f"=== DEBUG: Teacher dashboard POST request received ===")
        print(f"POST data keys: {list(request.POST.keys())}")
        print(f"FILES data keys: {list(request.FILES.keys())}")
        
        text = request.POST.get('status_text', '').strip()
        image = request.FILES.get('status_image')
        print(f"Text received: '{text}'")
        print(f"Image received: {image}")
        
        if text or image:
            try:
                status = StatusUpdate.objects.create(user=request.user, text=text, image=image)
                print(f"=== DEBUG: StatusUpdate created successfully with ID: {status.id} ===")
            except Exception as e:
                print(f"=== DEBUG: Error creating StatusUpdate: {e} ===")
        else:
            print(f"=== DEBUG: No text or image provided ===")
        return redirect('teacher_dashboard')
    
    # Handle homework creation
    if request.method == 'POST' and request.POST.get('form_type') == 'homework':
        title = request.POST.get('title')
        description = request.POST.get('description')
        assigned_class = request.POST.get('assigned_class')
        due_date = request.POST.get('due_date')
        file = request.FILES.get('file')
        
        if title and description and assigned_class and due_date:
            homework = Homework.objects.create(
                title=title,
                description=description,
                assigned_class=assigned_class,
                due_date=due_date,
                file=file
            )
            return redirect('teacher_dashboard')
    
    context = {
        'teacher': teacher,
        'total_students': total_students,
        'total_homeworks': total_homeworks,
        'total_posts': total_posts,
        'attendance_today': attendance_today,
        'posts': posts,
        'homeworks': homeworks,
        'students': students,
        'today': today,
        'user': request.user,
    }
    
    return render(request, 'teacher_dashboard.html', context)

def profile_page(request):
    if not request.user.is_authenticated:
        return redirect('login_page')
    context = {
        'student': None,
        'teacher': None,
        'user': request.user,
        'profile': getattr(request.user, 'profile', None),
    }
    profile = context['profile']
    if profile:
        if profile.role == 'student':
            student = getattr(request.user, 'student_profile', None)
            context['student'] = student
        elif profile.role == 'teacher':
            teacher = getattr(profile, 'teacher', None)
            context['teacher'] = teacher
    return render(request, 'profile.html', context)

def profile_edit(request):
    if not request.user.is_authenticated:
        return redirect('login_page')
    user = request.user
    profile = getattr(user, 'profile', None)
    if not profile or profile.role != 'student':
        messages.error(request, 'Only students can edit their profile.')
        return redirect('profile_page')
    student = getattr(user, 'student_profile', None)
    if request.method == 'POST':
        # Update fields from POST data
        student.full_name = request.POST.get('full_name', student.full_name)
        student.student_class = request.POST.get('student_class', student.student_class)
        student.symbol_number = request.POST.get('symbol_number', student.symbol_number)
        student.gender = request.POST.get('gender', student.gender)
        student.date_of_birth = request.POST.get('date_of_birth', student.date_of_birth)
        student.parent_name = request.POST.get('parent_name', student.parent_name)
        student.parent_contact = request.POST.get('parent_contact', student.parent_contact)
        student.phone_number = request.POST.get('phone_number', student.phone_number)
        student.address = request.POST.get('address', student.address)
        # Optionally handle profile photo upload
        if request.FILES.get('profile_photo'):
            student.profile_photo = request.FILES['profile_photo']
        if request.FILES.get('cover_photo'):
            student.cover_photo = request.FILES['cover_photo']
        student.save()
        messages.success(request, 'Profile updated successfully!')
        return redirect('profile_page')
    return render(request, 'profile_edit.html', {
        'student': student,
        'user': user,
    })

def teacher_login_view(request):
    import sys
    username = ''
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        print(f"[DEBUG] Received username: '{username}', password: '{password}'", file=sys.stderr)
        all_teachers = Teacher.objects.all()
        for t in all_teachers:
            print(f"[DEBUG] Teacher: username='{t.username}', password='{t.password}'", file=sys.stderr)
        try:
            teacher = Teacher.objects.get(username=username, password=password)
            # Use Django's login if teacher is linked to a user
            if teacher.profile and teacher.profile.user:
                login(request, teacher.profile.user)
            request.session['teacher_id'] = teacher.id
            request.session['role'] = 'teacher'  # Set role for the session
            return redirect('index')
        except Teacher.DoesNotExist:
            messages.error(request, 'Invalid username or password.')
    return render(request, 'teacher_login.html', {'username': username})

def logout_view(request):
    logout(request)
    return redirect('login_page')

# Create your views here.

class TeacherList(generics.ListAPIView):
    queryset = Teacher.objects.all()
    serializer_class = TeacherSerializer

class GalleryList(generics.ListAPIView):
    queryset = Gallery.objects.all()
    serializer_class = GallerySerializer

class HomeworkList(generics.ListAPIView):
    serializer_class = HomeworkSerializer

    def get_queryset(self):
        assigned_class = self.request.query_params.get('class')
        if assigned_class:
            return Homework.objects.filter(assigned_class=assigned_class)
        return Homework.objects.all()

class ContactMessageCreate(generics.CreateAPIView):
    queryset = ContactMessage.objects.all()
    serializer_class = ContactMessageSerializer

class ResultSearch(generics.ListAPIView):
    serializer_class = ResultSerializer

    def get_queryset(self):
        roll = self.request.query_params.get('roll_number')
        student_class = self.request.query_params.get('class')
        exam_type = self.request.query_params.get('exam_type')
        academic_year = self.request.query_params.get('academic_year')
        
        qs = Result.objects.filter(is_published=True)  # Only show published results
        
        if roll:
            qs = qs.filter(roll_number=roll)
        if student_class:
            qs = qs.filter(student_class=student_class)
        if exam_type:
            qs = qs.filter(exam_type=exam_type)
        if academic_year:
            qs = qs.filter(academic_year=academic_year)
            
        return qs.order_by('-uploaded_at')

# New enhanced results views
class ResultAnalyticsView(APIView):
    """Get analytics for results"""
    def get(self, request):
        class_name = request.query_params.get('class_name')
        exam_type = request.query_params.get('exam_type')
        academic_year = request.query_params.get('academic_year', '2024-25')
        
        from .models import ResultAnalytics
        analytics = ResultAnalytics.objects.all()
        
        if class_name:
            analytics = analytics.filter(class_name=class_name)
        if exam_type:
            analytics = analytics.filter(exam_type=exam_type)
        if academic_year:
            analytics = analytics.filter(academic_year=academic_year)
            
        data = []
        for analytic in analytics:
            data.append({
                'class_name': analytic.class_name,
                'exam_type': analytic.exam_type,
                'academic_year': analytic.academic_year,
                'total_students': analytic.total_students,
                'passed_students': analytic.passed_students,
                'failed_students': analytic.failed_students,
                'average_gpa': round(analytic.average_gpa, 2),
                'highest_gpa': analytic.highest_gpa,
                'lowest_gpa': analytic.lowest_gpa,
                'pass_percentage': round(analytic.pass_percentage, 2)
            })
            
        return Response(data)

class SubjectPerformanceView(APIView):
    """Get subject-wise performance"""
    def get(self, request):
        class_name = request.query_params.get('class_name')
        exam_type = request.query_params.get('exam_type')
        academic_year = request.query_params.get('academic_year', '2024-25')
        
        from .models import SubjectPerformance
        performances = SubjectPerformance.objects.all()
        
        if class_name:
            performances = performances.filter(class_name=class_name)
        if exam_type:
            performances = performances.filter(exam_type=exam_type)
        if academic_year:
            performances = performances.filter(academic_year=academic_year)
            
        data = []
        for performance in performances:
            data.append({
                'subject_name': performance.subject_name,
                'class_name': performance.class_name,
                'exam_type': performance.exam_type,
                'academic_year': performance.academic_year,
                'total_students': performance.total_students,
                'passed_students': performance.passed_students,
                'failed_students': performance.failed_students,
                'average_marks': round(performance.average_marks, 2),
                'highest_marks': performance.highest_marks,
                'lowest_marks': performance.lowest_marks,
                'pass_percentage': round(performance.pass_percentage, 2)
            })
            
        return Response(data)

class StudentPerformanceHistoryView(APIView):
    """Get student performance history"""
    def get(self, request):
        student_id = request.query_params.get('student_id')
        exam_type = request.query_params.get('exam_type')
        academic_year = request.query_params.get('academic_year')
        
        if not student_id:
            return Response({'error': 'student_id is required'}, status=400)
            
        try:
            student = Student.objects.get(id=student_id)
        except Student.DoesNotExist:
            return Response({'error': 'Student not found'}, status=404)
            
        from .models import StudentPerformanceHistory
        history = StudentPerformanceHistory.objects.filter(student=student)
        
        if exam_type:
            history = history.filter(exam_type=exam_type)
        if academic_year:
            history = history.filter(academic_year=academic_year)
            
        data = []
        for record in history:
            data.append({
                'exam_type': record.exam_type,
                'academic_year': record.academic_year,
                'class_name': record.class_name,
                'gpa': record.gpa,
                'total_marks': record.total_marks,
                'class_position': record.class_position,
                'total_students': record.total_students,
                'performance_status': record.performance_status,
                'created_at': record.created_at
            })
            
        return Response(data)

class ResultTemplateView(APIView):
    """Get result templates"""
    def get(self, request):
        class_name = request.query_params.get('class_name')
        exam_type = request.query_params.get('exam_type')
        
        templates = ResultTemplate.objects.filter(is_active=True)
        
        if class_name:
            templates = templates.filter(class_name=class_name)
        if exam_type:
            templates = templates.filter(exam_type=exam_type)
            
        data = []
        for template in templates:
            data.append({
                'id': template.id,
                'name': template.name,
                'exam_type': template.exam_type,
                'class_name': template.class_name,
                'subjects': template.subjects,
                'created_at': template.created_at
            })
            
        return Response(data)

class GradeScaleView(APIView):
    """Get grade scales"""
    def get(self, request):
        class_name = request.query_params.get('class_name')
        
        scales = GradeScale.objects.all()
        
        if class_name:
            scales = scales.filter(class_name=class_name)
            
        data = []
        for scale in scales:
            data.append({
                'class_name': scale.class_name,
                'grade': scale.grade,
                'min_marks': scale.min_marks,
                'max_marks': scale.max_marks,
                'grade_point': scale.grade_point,
                'description': scale.description,
                'is_pass': scale.is_pass
            })
            
        return Response(data)

class NoticeList(generics.ListAPIView):
    queryset = Notice.objects.all().order_by('-date')
    serializer_class = NoticeSerializer

class ResourceList(generics.ListAPIView):
    """Public list of published resources with filters: class, subject, query, type."""
    serializer_class = ResourceSerializer
    permission_classes = [permissions.AllowAny]

    def get_queryset(self):
        qs = Resource.objects.filter(is_published=True)
        class_name = self.request.query_params.get('class')
        subject_id = self.request.query_params.get('subject')
        query = self.request.query_params.get('q')
        is_qp = self.request.query_params.get('question_paper')
        if class_name:
            qs = qs.filter(class_name=class_name)
        if subject_id:
            qs = qs.filter(subject_id=subject_id)
        if query:
            qs = qs.filter(models.Q(title__icontains=query) | models.Q(description__icontains=query) | models.Q(tags__icontains=query))
        if is_qp in ['true','1','yes']:
            qs = qs.filter(is_question_paper=True)
        return qs.order_by('-created_at')

class StudentListCreate(generics.ListCreateAPIView):
    queryset = Student.objects.all().order_by('-created_at')
    serializer_class = StudentSerializer
    permission_classes = [permissions.IsAdminUser]

    def get_queryset(self):
        queryset = super().get_queryset()
        class_name = self.request.query_params.get('class')
        if class_name:
            queryset = queryset.filter(student_class=class_name)
        return queryset

class StudentRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    permission_classes = [permissions.IsAdminUser]

class StudentSignup(generics.CreateAPIView):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    permission_classes = [permissions.AllowAny]

class LoginView(APIView):
    def post(self, request):
        email = request.data.get('email')
        password = request.data.get('password')
        user = authenticate(request, username=email, password=password)
        if user is not None:
            login(request, user)
            return Response({'success': True, 'user_id': user.id, 'username': user.username})
        else:
            return Response({'success': False, 'error': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)

class LogoutView(APIView):
    def post(self, request):
        logout(request)
        return Response({'success': True})

class RegisterView(APIView):
    def post(self, request):
        serializer = RegisterSerializer(data=request.data)
        if serializer.is_valid():
            obj = serializer.save()
            role = request.data.get('role')
            return Response({'success': True, 'role': role}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class UserRoleView(APIView):
    def get(self, request):
        if not request.user.is_authenticated:
            return Response({'role': None})
        try:
            role = request.user.profile.role
        except Exception:
            role = None
        return Response({'role': role})

class StudentListAll(generics.ListAPIView):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    permission_classes = [permissions.IsAdminUser]

class TeacherListAll(generics.ListAPIView):
    queryset = Teacher.objects.all()
    serializer_class = TeacherSerializer
    permission_classes = [permissions.IsAdminUser]

class HomeworkDueTodayView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        user = request.user
        try:
            student = Student.objects.get(user=user)
        except Student.DoesNotExist:
            return Response({'error': 'Student not found.'}, status=404)
        today = timezone.now().date()
        homeworks = Homework.objects.filter(assigned_class=student.student_class, due_date=today)
        serializer = HomeworkSerializer(homeworks, many=True)
        return Response(serializer.data)

class RecentSubmissionsView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        user = request.user
        try:
            student = Student.objects.get(user=user)
        except Student.DoesNotExist:
            return Response({'error': 'Student not found.'}, status=404)
        submissions = HomeworkSubmission.objects.filter(student=student).order_by('-submitted_at')[:5]
        serializer = HomeworkSubmissionSerializer(submissions, many=True)
        return Response(serializer.data)

class HomeworkSubmissionCreateView(generics.CreateAPIView):
    serializer_class = HomeworkSubmissionSerializer
    permission_classes = [IsAuthenticated]
    def perform_create(self, serializer):
        user = self.request.user
        try:
            student = Student.objects.get(user=user)
        except Student.DoesNotExist:
            raise serializers.ValidationError('Student not found.')
        serializer.save(student=student)

class ActivityLogView(generics.CreateAPIView):
    queryset = ActivityLog.objects.all()
    serializer_class = ActivityLogSerializer
    permission_classes = [permissions.AllowAny]

    def perform_create(self, serializer):
        user = self.request.user if self.request.user.is_authenticated else None
        serializer.save(user=user)

def post_homework_view(request):
    teacher_id = request.session.get('teacher_id')
    if not teacher_id:
        return redirect('teacher_login')
    if request.method == 'POST':
        form = HomeworkForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Homework posted successfully!')
            return redirect('teacher_dashboard')
    else:
        form = HomeworkForm()
    return render(request, 'post_homework.html', {'form': form})

def view_submissions_view(request):
    teacher_id = request.session.get('teacher_id')
    if not teacher_id:
        return redirect('teacher_login')
    # For now, show all submissions for all homework
    from .models import HomeworkSubmission
    submissions = HomeworkSubmission.objects.select_related('homework', 'student').order_by('-submitted_at')
    return render(request, 'view_submissions.html', {'submissions': submissions})

def teacher_attendance_view(request):
    teacher_id = request.session.get('teacher_id')
    if not teacher_id:
        return redirect('teacher_login')
    from .models import Teacher, Student, Attendance, ClassSection
    import datetime
    teacher = Teacher.objects.get(id=teacher_id)
    today = timezone.localdate()
    # Get all class sections for dropdown
    class_sections = ClassSection.objects.all().order_by('name')
    # Get selected class and date from GET or POST
    selected_class_id = request.GET.get('class') or request.POST.get('class_section')
    selected_date_str = request.GET.get('date') or request.POST.get('date')
    selected_date = today
    if selected_date_str:
        try:
            selected_date = datetime.datetime.strptime(selected_date_str, '%Y-%m-%d').date()
        except Exception:
            selected_date = today
    students = Student.objects.none()
    selected_class = None
    if selected_class_id:
        try:
            selected_class = ClassSection.objects.get(id=selected_class_id)
            students = Student.objects.filter(class_section=selected_class).order_by('full_name')
        except ClassSection.DoesNotExist:
            students = Student.objects.none()
    message = ''
    locked = False
    # Check if attendance is locked (after 24h or locked flag)
    if selected_class:
        any_attendance = Attendance.objects.filter(class_section=selected_class, date=selected_date)
        if any_attendance.exists():
            first = any_attendance.first()
            locked = first.locked or (timezone.now().date() > selected_date and (timezone.now() - first.created_at).total_seconds() > 86400)
    if request.method == 'POST' and students and not locked:
        for student in students:
            status = request.POST.get(f'status_{student.id}')
            remarks = request.POST.get(f'remarks_{student.id}', '')
            if status:
                att, created = Attendance.objects.update_or_create(
                    student=student,
                    class_section=selected_class,
                    date=selected_date,
                    defaults={'status': status, 'marked_by': teacher, 'remarks': remarks}
                )
        # Lock attendance if date is in the past
        if selected_date < today:
            Attendance.objects.filter(class_section=selected_class, date=selected_date).update(locked=True)
        message = 'Attendance marked successfully!'
    # Get attendance records for this class/date
    attendance_records = {a.student_id: a for a in Attendance.objects.filter(class_section=selected_class, date=selected_date)} if selected_class else {}
    return render(request, 'teacher_attendance.html', {
        'teacher': teacher,
        'students': students,
        'attendance_records': attendance_records,
        'today': today,
        'selected_date': selected_date,
        'class_sections': class_sections,
        'selected_class': selected_class,
        'locked': locked,
        'message': message,
    })

def attendance_overview_view(request):
    from .models import Attendance, Student, ClassSection
    import datetime
    today = timezone.localdate()
    class_sections = ClassSection.objects.all().order_by('name')
    selected_class_id = request.GET.get('class')
    selected_date_str = request.GET.get('date')
    selected_date = today
    if selected_date_str:
        try:
            selected_date = datetime.datetime.strptime(selected_date_str, '%Y-%m-%d').date()
        except Exception:
            selected_date = today
    students = Student.objects.all().order_by('full_name')
    selected_class = None
    if selected_class_id:
        try:
            selected_class = ClassSection.objects.get(id=selected_class_id)
            students = students.filter(class_section=selected_class)
        except ClassSection.DoesNotExist:
            students = Student.objects.none()
    attendance_records = {a.student_id: a for a in Attendance.objects.filter(date=selected_date)}
    # Summary stats
    summary = {'present': 0, 'absent': 0, 'late': 0, 'leave': 0, 'excused': 0}
    for att in attendance_records.values():
        summary[att.status] = summary.get(att.status, 0) + 1
    total = len(students)
    return render(request, 'attendance_overview.html', {
        'students': students,
        'attendance_records': attendance_records,
        'today': today,
        'selected_date': selected_date,
        'class_sections': class_sections,
        'selected_class': selected_class,
        'summary': summary,
        'total': total,
    })

@csrf_exempt
def gallery_likes_api(request, pk):
    from .models import GalleryLike
    user = request.user if request.user.is_authenticated else None
    session_key = request.session.session_key or ''
    if not session_key:
        request.session.save()
        session_key = request.session.session_key

    if request.method == 'GET':
        likes = GalleryLike.objects.filter(gallery_id=pk)
        counts = {r: likes.filter(reaction=r).count() for r, _ in GalleryLike.REACTION_CHOICES}
        # Determine if current user/session has liked
        if user:
            liked = GalleryLike.objects.filter(gallery_id=pk, user=user, reaction='like').exists()
        else:
            liked = GalleryLike.objects.filter(gallery_id=pk, session_key=session_key, reaction='like').exists()
        return JsonResponse({'counts': counts, 'liked': liked})

    if request.method == 'POST':
        import json
        data = json.loads(request.body)
        reaction = data.get('reaction', 'like')
        if user:
            like = GalleryLike.objects.filter(gallery_id=pk, user=user, reaction='like').first()
            if like:
                like.delete()  # Unlike
                liked = False
            else:
                GalleryLike.objects.create(
                    gallery_id=pk,
                    user=user,
                    ip_address=request.META.get('REMOTE_ADDR'),
                    reaction='like',
                    session_key=session_key
                )
                liked = True
        else:
            like = GalleryLike.objects.filter(gallery_id=pk, session_key=session_key, reaction='like').first()
            if like:
                like.delete()  # Unlike
                liked = False
            else:
                GalleryLike.objects.create(
                    gallery_id=pk,
                    session_key=session_key,
                    ip_address=request.META.get('REMOTE_ADDR'),
                    reaction='like'
                )
                liked = True
        likes = GalleryLike.objects.filter(gallery_id=pk)
        counts = {r: likes.filter(reaction=r).count() for r, _ in GalleryLike.REACTION_CHOICES}
        return JsonResponse({'reacted': True, 'counts': counts, 'liked': liked})

@csrf_exempt
def gallery_comments_api(request, pk):
    from .models import GalleryComment, Student, Teacher
    if request.method == 'GET':
        comments = GalleryComment.objects.filter(gallery_id=pk).order_by('-created_at')
        data = []
        for c in comments:
            # Try to get profile pic from student or teacher
            profile_pic = None
            if c.student and c.student.profile_photo:
                profile_pic = c.student.profile_photo.url
            elif c.teacher and c.teacher.photo:
                profile_pic = c.teacher.photo.url
            data.append({
                'name': c.name,
                'comment': c.comment,
                'created_at': c.created_at.strftime('%Y-%m-%d %H:%M'),
                'profile_pic': profile_pic
            })
        return JsonResponse({'comments': data})
    if request.method == 'POST':
        import json
        data = json.loads(request.body)
        # Always use logged-in user's profile name if authenticated
        user = request.user if request.user.is_authenticated else None
        student = None
        teacher = None
        name = None
        if user:
            student = Student.objects.filter(user=user).first()
            if student:
                name = student.full_name
            else:
                teacher = Teacher.objects.filter(profile__user=user).first()
                if teacher:
                    name = teacher.full_name
            if not name:
                name = user.get_full_name() or user.username
        else:
            name = data.get('name', 'Anonymous')
        comment = GalleryComment.objects.create(
            gallery_id=pk,
            name=name,
            email=data.get('email', ''),
            comment=data.get('comment', ''),
            user=user,
            student=student,
            teacher=teacher
        )
        # Return profile_pic in response
        profile_pic = None
        if student and student.profile_photo:
            profile_pic = student.profile_photo.url
        elif teacher and teacher.photo:
            profile_pic = teacher.photo.url
        return JsonResponse({'success': True, 'comment': {'name': comment.name, 'comment': comment.comment, 'created_at': comment.created_at.strftime('%Y-%m-%d %H:%M'), 'profile_pic': profile_pic}})

@csrf_exempt
def post_like_api(request, pk):
    from .models import Post
    user = request.user if request.user.is_authenticated else None
    post = Post.objects.get(pk=pk)
    if request.method == 'POST' and user:
        if post.likes.filter(id=user.id).exists():
            post.likes.remove(user)
            liked = False
        else:
            post.likes.add(user)
            liked = True
        return JsonResponse({'liked': liked, 'count': post.likes.count()})
    return JsonResponse({'liked': user and post.likes.filter(id=user.id).exists(), 'count': post.likes.count()})

@csrf_exempt
def post_comments_api(request, pk):
    from .models import Post, PostComment
    user = request.user if request.user.is_authenticated else None
    post = Post.objects.get(pk=pk)
    if request.method == 'POST' and user:
        import json
        data = json.loads(request.body)
        text = data.get('text', '').strip()
        if text:
            comment = PostComment.objects.create(post=post, user=user, text=text)
            return JsonResponse({'success': True, 'comment': {'user': user.get_full_name() or user.username, 'text': comment.text, 'created_at': comment.created_at.strftime('%Y-%m-%d %H:%M')}})
        return JsonResponse({'success': False, 'error': 'Empty comment'})
    # GET: return comments
    comments = post.comments.select_related('user').order_by('-created_at')
    data = [{'user': c.user.get_full_name() or c.user.username, 'text': c.text, 'created_at': c.created_at.strftime('%Y-%m-%d %H:%M')} for c in comments]
    return JsonResponse({'comments': data})

@csrf_exempt
def status_like_api(request, pk):
    from .models import StatusUpdate, StatusLike
    user = request.user if request.user.is_authenticated else None
    status = StatusUpdate.objects.get(pk=pk)
    if request.method == 'POST' and user:
        like, created = StatusLike.objects.get_or_create(status=status, user=user)
        if not created:
            like.delete()
            liked = False
        else:
            liked = True
        return JsonResponse({'liked': liked, 'count': status.likes.count()})
    return JsonResponse({'liked': user and status.likes.filter(user=user).exists(), 'count': status.likes.count()})

@csrf_exempt
def status_comments_api(request, pk):
    from .models import StatusUpdate, StatusComment
    user = request.user if request.user.is_authenticated else None
    status = StatusUpdate.objects.get(pk=pk)
    if request.method == 'POST' and user:
        import json
        data = json.loads(request.body)
        text = data.get('text', '').strip()
        if text:
            comment = StatusComment.objects.create(status=status, user=user, text=text)
            return JsonResponse({'success': True, 'comment': {'user': user.get_full_name() or user.username, 'text': comment.text, 'created_at': comment.created_at.strftime('%Y-%m-%d %H:%M')}})
        return JsonResponse({'success': False, 'error': 'Empty comment'})
    # GET: return comments
    comments = status.comments.select_related('user').order_by('-created_at')
    data = [{'user': c.user.get_full_name() or c.user.username, 'text': c.text, 'created_at': c.created_at.strftime('%Y-%m-%d %H:%M')} for c in comments]
    return JsonResponse({'comments': data})

@csrf_exempt
def delete_status_api(request, pk):
    """Delete a status update (post) - only creator can delete"""
    from .models import StatusUpdate
    user = request.user if request.user.is_authenticated else None
    
    if not user:
        return JsonResponse({'success': False, 'error': 'Authentication required'})
    
    try:
        status = StatusUpdate.objects.get(pk=pk)
        
        # Check if user is the creator
        if status.user != user:
            return JsonResponse({'success': False, 'error': 'You can only delete your own posts'})
        
        # Delete the status
        status.delete()
        return JsonResponse({'success': True, 'message': 'Post deleted successfully'})
        
    except StatusUpdate.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Post not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@csrf_exempt
def delete_comment_api(request, comment_id):
    """Delete a comment - only creator can delete"""
    from .models import StatusComment
    user = request.user if request.user.is_authenticated else None
    
    if not user:
        return JsonResponse({'success': False, 'error': 'Authentication required'})
    
    try:
        comment = StatusComment.objects.get(pk=comment_id)
        
        # Check if user is the creator
        if comment.user != user:
            return JsonResponse({'success': False, 'error': 'You can only delete your own comments'})
        
        # Delete the comment
        comment.delete()
        return JsonResponse({'success': True, 'message': 'Comment deleted successfully'})
        
    except StatusComment.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Comment not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@csrf_exempt
def delete_homework_api(request, homework_id):
    """Delete a homework - only teachers can delete"""
    from .models import Homework
    user = request.user if request.user.is_authenticated else None
    
    if not user:
        return JsonResponse({'success': False, 'error': 'Authentication required'})
    
    # Check if user is a teacher
    if not hasattr(user, 'teacher_profile'):
        return JsonResponse({'success': False, 'error': 'Only teachers can delete homework'})
    
    try:
        homework = Homework.objects.get(pk=homework_id)
        
        # Delete the homework
        homework.delete()
        return JsonResponse({'success': True, 'message': 'Homework deleted successfully'})
        
    except Homework.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Homework not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@csrf_exempt
def follow_user_api(request, username):
    """Follow or unfollow a user"""
    from .models import Follow
    from django.contrib.auth.models import User
    
    user = request.user if request.user.is_authenticated else None
    
    if not user:
        return JsonResponse({'success': False, 'error': 'Authentication required'})
    
    if user.username == username:
        return JsonResponse({'success': False, 'error': 'You cannot follow yourself'})
    
    try:
        target_user = User.objects.get(username=username)
        
        # Check if already following
        follow_relation, created = Follow.objects.get_or_create(
            follower=user,
            following=target_user
        )
        
        if created:
            # New follow relationship created
            return JsonResponse({
                'success': True, 
                'is_following': True, 
                'message': f'You are now following {username}'
            })
        else:
            # Already following, so unfollow
            follow_relation.delete()
            return JsonResponse({
                'success': True, 
                'is_following': False, 
                'message': f'You unfollowed {username}'
            })
            
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'User not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@csrf_exempt
def toggle_follow_api(request):
    """Toggle follow/unfollow a user by user_id"""
    from .models import Follow
    from django.contrib.auth.models import User
    import json
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Only POST method is allowed'})
    
    user = request.user if request.user.is_authenticated else None
    
    if not user:
        return JsonResponse({'success': False, 'error': 'Authentication required'})
    
    try:
        data = json.loads(request.body)
        user_id = data.get('user_id')
        
        if not user_id:
            return JsonResponse({'success': False, 'error': 'User ID is required'})
        
        target_user = User.objects.get(id=user_id)
        
        if user.id == target_user.id:
            return JsonResponse({'success': False, 'error': 'You cannot follow yourself'})
        
        # Check if already following
        follow_relation = Follow.objects.filter(follower=user, following=target_user).first()
        
        if follow_relation:
            # Already following, so unfollow
            follow_relation.delete()
            return JsonResponse({
                'success': True, 
                'is_following': False, 
                'message': f'You unfollowed {target_user.username}'
            })
        else:
            # Not following, so follow
            Follow.objects.create(follower=user, following=target_user)
            return JsonResponse({
                'success': True, 
                'is_following': True, 
                'message': f'You are now following {target_user.username}'
            })
            
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'User not found'})
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@csrf_exempt
def friend_suggestions_api(request):
    """Get friend suggestions for the current user"""
    from .models import Follow
    from django.contrib.auth.models import User
    from django.db.models import Count, Q
    
    user = request.user if request.user.is_authenticated else None
    
    if not user:
        return JsonResponse({'success': False, 'error': 'Authentication required'})
    
    try:
        # Get users the current user is following
        following_ids = user.following.values_list('following_id', flat=True)
        
        # Get users who are followed by users the current user is following (friends of friends)
        friends_of_friends = Follow.objects.filter(
            follower_id__in=following_ids
        ).exclude(
            following=user  # Exclude the current user
        ).exclude(
            following_id__in=following_ids  # Exclude users the current user is already following
        ).values('following_id').annotate(
            common_friends=Count('following_id')
        ).order_by('-common_friends')[:10]
        
        friend_suggestion_ids = [f['following_id'] for f in friends_of_friends]
        
        # If we don't have enough suggestions, add some random users
        if len(friend_suggestion_ids) < 5:
            # Get random users not already in suggestions and not already followed
            random_users = User.objects.exclude(
                Q(id=user.id) | Q(id__in=following_ids) | Q(id__in=friend_suggestion_ids)
            ).order_by('?')[:5-len(friend_suggestion_ids)]
            
            friend_suggestion_ids.extend([u.id for u in random_users])
        
        # Get the actual user objects
        suggestions = User.objects.filter(id__in=friend_suggestion_ids)
        
        # Format the response
        result = []
        for suggestion in suggestions:
            profile_photo = None
            if hasattr(suggestion, 'student_profile') and suggestion.student_profile and suggestion.student_profile.profile_photo:
                profile_photo = suggestion.student_profile.profile_photo.url
            elif hasattr(suggestion, 'teacher_profile') and suggestion.teacher_profile and suggestion.teacher_profile.photo:
                profile_photo = suggestion.teacher_profile.photo.url
                
            result.append({
                'id': suggestion.id,
                'username': suggestion.username,
                'full_name': suggestion.get_full_name() or suggestion.username,
                'profile_photo': profile_photo
            })
        
        return JsonResponse(result, safe=False)
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@csrf_exempt
def search_friends_api(request):
    """Search for users by name or username"""
    from .models import Follow
    from django.contrib.auth.models import User
    from django.db.models import Q
    
    user = request.user if request.user.is_authenticated else None
    
    if not user:
        return JsonResponse({'success': False, 'error': 'Authentication required'})
    
    try:
        query = request.GET.get('q', '').strip()
        
        if not query:
            return JsonResponse([], safe=False)
        
        # Search for users by first name, last name, or username
        users = User.objects.filter(
            Q(first_name__icontains=query) | 
            Q(last_name__icontains=query) | 
            Q(username__icontains=query)
        ).exclude(id=user.id)[:10]  # Exclude current user and limit to 10 results
        
        # Get the IDs of users the current user is following
        following_ids = set(user.following.values_list('following_id', flat=True))
        
        # Format the response
        result = []
        for found_user in users:
            profile_photo = None
            if hasattr(found_user, 'student_profile') and found_user.student_profile and found_user.student_profile.profile_photo:
                profile_photo = found_user.student_profile.profile_photo.url
            elif hasattr(found_user, 'teacher_profile') and found_user.teacher_profile and found_user.teacher_profile.photo:
                profile_photo = found_user.teacher_profile.photo.url
                
            result.append({
                'id': found_user.id,
                'username': found_user.username,
                'full_name': found_user.get_full_name() or found_user.username,
                'profile_photo': profile_photo,
                'is_following': found_user.id in following_ids
            })
        
        return JsonResponse(result, safe=False)
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def view_profile(request, username):
    """View a user's profile"""
    from django.contrib.auth.models import User
    from .models import Follow, StatusUpdate
    
    try:
        profile_user = User.objects.get(username=username)
        
        # Get user's posts
        posts = StatusUpdate.objects.filter(user=profile_user).order_by('-created_at')
        
        # Get follow statistics
        followers_count = profile_user.followers.count()
        following_count = profile_user.following.count()
        
        # Check if current user is following this user
        is_following = False
        if request.user.is_authenticated:
            is_following = Follow.objects.filter(follower=request.user, following=profile_user).exists()
        
        # Get user's profile information
        profile_data = {
            'username': profile_user.username,
            'full_name': profile_user.get_full_name() or profile_user.username,
            'email': profile_user.email,
            'date_joined': profile_user.date_joined,
            'posts_count': posts.count(),
            'followers_count': followers_count,
            'following_count': following_count,
            'is_following': is_following,
        }
        
        # Add profile picture
        if hasattr(profile_user, 'student_profile') and profile_user.student_profile.profile_photo:
            profile_data['profile_picture'] = profile_user.student_profile.profile_photo.url
        elif hasattr(profile_user, 'teacher_profile') and profile_user.teacher_profile.photo:
            profile_data['profile_picture'] = profile_user.teacher_profile.photo.url
        else:
            profile_data['profile_picture'] = None
        
        # Add cover photo
        if hasattr(profile_user, 'student_profile') and profile_user.student_profile.cover_photo:
            profile_data['cover_photo'] = profile_user.student_profile.cover_photo.url
        elif hasattr(profile_user, 'teacher_profile') and profile_user.teacher_profile.cover_photo:
            profile_data['cover_photo'] = profile_user.teacher_profile.cover_photo.url
        else:
            profile_data['cover_photo'] = None
        
        context = {
            'profile_user': profile_user,
            'profile_data': profile_data,
            'posts': posts[:10],  # Show last 10 posts
            'user': request.user,
        }
        
        return render(request, 'view_profile.html', context)
        
    except User.DoesNotExist:
        return render(request, '404.html', {'message': 'User not found'}, status=404)

def unified_dashboard_view(request):
    if not request.user.is_authenticated:
        return redirect('login_page')
    profile = getattr(request.user, 'profile', None)
    if profile:
        if profile.role == 'student':
            return redirect('student_feed')  # Redirect to /feed/ for students
        elif profile.role == 'teacher':
            return redirect('teacher_dashboard')
    # Fallback: if no profile or unknown role, go to profile page
    return redirect('profile_page')

def my_details_view(request):
    if not request.user.is_authenticated:
        return redirect('login_page')
    context = {
        'student': None,
        'teacher': None,
        'user': request.user,
        'profile': getattr(request.user, 'profile', None),
    }
    profile = context['profile']
    if profile:
        if profile.role == 'student':
            context['student'] = getattr(request.user, 'student_profile', None)
        elif profile.role == 'teacher':
            context['teacher'] = getattr(profile, 'teacher', None)
    return render(request, 'my_details.html', context)

def student_feed(request):
    from .models import StatusUpdate
    
    print(f"=== DEBUG: student_feed view called ===")
    print(f"Request method: {request.method}")
    print(f"User authenticated: {request.user.is_authenticated}")
    if request.user.is_authenticated:
        print(f"User: {request.user.username}")
    
    # Add a simple test for any POST request
    if request.method == 'POST':
        print(f"=== DEBUG: POST request received ===")
        print(f"POST data keys: {list(request.POST.keys())}")
        print(f"FILES data keys: {list(request.FILES.keys())}")
        print(f"All POST data: {dict(request.POST)}")
    
    if request.method == 'POST' and request.user.is_authenticated and 'status_text' in request.POST:
        print(f"=== DEBUG: Processing status_text POST request ===")
        text = request.POST.get('status_text', '').strip()
        image = request.FILES.get('status_image')
        print(f"Text received: '{text}'")
        print(f"Image received: {image}")
        
        if text or image:
            try:
                status = StatusUpdate.objects.create(user=request.user, text=text, image=image)
                print(f"=== DEBUG: StatusUpdate created successfully with ID: {status.id} ===")
            except Exception as e:
                print(f"=== DEBUG: Error creating StatusUpdate: {e} ===")
        else:
            print(f"=== DEBUG: No text or image provided ===")
        return redirect('student_feed')
    
    # Show all status updates from all users
    posts = StatusUpdate.objects.all().order_by('-created_at')
    
    return render(request, 'student_feed.html', {'posts': posts, 'user': request.user})

def public_profile_view(request, username):
    # Get the user by username
    user_obj = User.objects.filter(username=username).first()
    if not user_obj:
        return render(request, 'profile_not_found.html', {'username': username})
    # Handle follow/unfollow POST
    if request.method == 'POST' and request.user.is_authenticated and request.user != user_obj:
        from .models import Follow
        action = request.POST.get('action')
        if action == 'follow':
            Follow.objects.get_or_create(follower=request.user, following=user_obj)
        elif action == 'unfollow':
            Follow.objects.filter(follower=request.user, following=user_obj).delete()
        return redirect('public_profile', username=username)
    # Get student/teacher profile if exists
    student = getattr(user_obj, 'student_profile', None)
    teacher = getattr(user_obj, 'teacher_profile', None)
    # Get posts and statuses
    posts = user_obj.posts.all().order_by('-created_at') if hasattr(user_obj, 'posts') else []
    statuses = user_obj.status_updates.all().order_by('-created_at') if hasattr(user_obj, 'status_updates') else []
    # Followers and following
    followers = user_obj.followers.all()
    following = user_obj.following.all()
    is_following = False
    is_following_back = False
    is_friend = False
    if request.user.is_authenticated and request.user != user_obj:
        is_following = user_obj.followers.filter(follower=request.user).exists()
        is_following_back = user_obj.following.filter(following=request.user).exists()
        is_friend = is_following and is_following_back
    friends = []
    if request.user.is_authenticated:
        # Friends = users who both follow and are followed by profile_user
        follower_ids = set(f.follower_id for f in user_obj.followers.all())
        following_ids = set(f.following_id for f in user_obj.following.all())
        friend_ids = follower_ids & following_ids
        friends = User.objects.filter(id__in=friend_ids)
    return render(request, 'public_profile.html', {
        'profile_user': user_obj,
        'student': student,
        'teacher': teacher,
        'posts': posts,
        'statuses': statuses,
        'followers': followers,
        'following': following,
        'is_following': is_following,
        'is_following_back': is_following_back,
        'is_friend': is_friend,
        'user': request.user,
        'friends': friends,
    })

@require_GET
def next_symbol_number_api(request):
    from .models import Student
    last_student = Student.objects.order_by('-id').first()
    if last_student and last_student.symbol_number and last_student.symbol_number.isdigit():
        next_symbol = str(int(last_student.symbol_number) + 1)
    else:
        next_symbol = '1001'
    return JsonResponse({'next_symbol_number': next_symbol})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def teacher_dashboard_api(request):
    user = request.user
    if not hasattr(user, 'profile') or user.profile.role != 'teacher':
        return Response({'error': 'Not a teacher'}, status=403)
    teacher = Teacher.objects.filter(profile=user.profile).first()
    from .models import ClassSection, Student, Homework, HomeworkSubmission, Notice
    classes = ClassSection.objects.filter(teachers=teacher)
    num_classes = classes.count()
    students = Student.objects.filter(class_section__in=classes)
    num_students = students.count()
    homeworks = Homework.objects.filter(assigned_class__in=classes.values_list('name', flat=True))
    num_homeworks = homeworks.count()
    pending_submissions = HomeworkSubmission.objects.filter(homework__in=homeworks, status='pending')
    num_pending_submissions = pending_submissions.count()
    recent_submissions = HomeworkSubmission.objects.filter(homework__in=homeworks).order_by('-submitted_at')[:5]
    recent_notices = Notice.objects.order_by('-date')[:3]
    return Response({
        'teacher': {
            'full_name': teacher.full_name,
            'subject': teacher.subject,
            'photo': teacher.photo.url if teacher.photo else '',
        },
        'num_classes': num_classes,
        'num_students': num_students,
        'num_homeworks': num_homeworks,
        'num_pending_submissions': num_pending_submissions,
        'recent_submissions': [
            {
                'student': sub.student.full_name,
                'homework': sub.homework.title,
                'submitted_at': sub.submitted_at,
            } for sub in recent_submissions
        ],
        'recent_notices': [
            {
                'title': n.title,
                'date': n.date,
                'category': n.category,
            } for n in recent_notices
        ],
        'classes': [c.name for c in classes],
    })

@csrf_exempt
@require_http_methods(["POST"])
def import_results_excel(request):
    """Handle Excel/CSV import for results"""
    if not PANDAS_AVAILABLE:
        return JsonResponse({'success': False, 'error': 'Pandas is not installed. Please install pandas and openpyxl for Excel import functionality.'})
    
    try:
        if 'file' not in request.FILES:
            return JsonResponse({'success': False, 'error': 'No file uploaded'})
        
        file = request.FILES['file']
        
        # Check file type
        if file.name.endswith('.csv'):
            # Handle CSV
            content = file.read().decode('utf-8')
            df = pd.read_csv(io.StringIO(content))
        elif file.name.endswith(('.xlsx', '.xls')):
            # Handle Excel
            df = pd.read_excel(file)
        else:
            return JsonResponse({'success': False, 'error': 'Unsupported file format. Please upload CSV or Excel file.'})
        
        # Process the data
        results_data = []
        for index, row in df.iterrows():
            try:
                # Extract data from row
                student_name = str(row.get('Student Name', '')).strip()
                if not student_name or student_name == 'nan':
                    continue
                
                # Get subject marks
                nepali = float(row.get('Nepali', 0)) if pd.notna(row.get('Nepali')) else 0
                math = float(row.get('Math', 0)) if pd.notna(row.get('Math')) else 0
                science = float(row.get('Science', 0)) if pd.notna(row.get('Science')) else 0
                computer = float(row.get('Computer', 0)) if pd.notna(row.get('Computer')) else 0
                social = float(row.get('Social', 0)) if pd.notna(row.get('Social')) else 0
                
                # Calculate totals
                total = nepali + math + science + computer + social
                percentage = (total / 500) * 100 if total > 0 else 0
                
                # Calculate GPA
                if percentage >= 90:
                    gpa = 4.0
                elif percentage >= 80:
                    gpa = 3.6
                elif percentage >= 70:
                    gpa = 3.2
                elif percentage >= 60:
                    gpa = 2.8
                elif percentage >= 50:
                    gpa = 2.4
                elif percentage >= 40:
                    gpa = 2.0
                elif percentage >= 20:
                    gpa = 1.6
                else:
                    gpa = 0.8
                
                # Get other fields
                roll_number = str(row.get('Roll Number', f'R{index+1:03d}')).strip()
                student_class = str(row.get('Class', '10')).strip()
                exam_type = str(row.get('Exam Type', 'Final Term')).strip()
                academic_year = str(row.get('Academic Year', '2024-25')).strip()
                
                results_data.append({
                    'student_name': student_name,
                    'roll_number': roll_number,
                    'student_class': student_class,
                    'total': total,
                    'gpa': gpa,
                    'percentage': percentage,
                    'exam_type': exam_type,
                    'academic_year': academic_year,
                    'nepali': nepali,
                    'math': math,
                    'science': science,
                    'computer': computer,
                    'social': social
                })
                
            except Exception as e:
                continue  # Skip problematic rows
        
        return JsonResponse({
            'success': True,
            'data': results_data,
            'message': f'Successfully processed {len(results_data)} results'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["GET"])
def export_results_excel(request):
    """Export results to Excel/CSV"""
    if not PANDAS_AVAILABLE:
        return JsonResponse({'success': False, 'error': 'Pandas is not installed. Please install pandas and openpyxl for Excel export functionality.'})
    
    try:
        from .models import Result
        
        # Get results based on filters
        class_filter = request.GET.get('class', '')
        exam_filter = request.GET.get('exam', '')
        year_filter = request.GET.get('year', '')
        
        results = Result.objects.all()
        
        if class_filter:
            results = results.filter(student_class=class_filter)
        if exam_filter:
            results = results.filter(exam_type=exam_filter)
        if year_filter:
            results = results.filter(academic_year=year_filter)
        
        # Create DataFrame
        data = []
        for result in results:
            data.append({
                'Student Name': result.student_name,
                'Roll Number': result.roll_number,
                'Class': result.student_class,
                'Total': result.total,
                'Percentage': result.percentage,
                'GPA': result.gpa,
                'Grade': result.get_performance_status(),
                'Remarks': result.remarks or '',
                'Exam Type': result.exam_type,
                'Academic Year': result.academic_year,
                'Published': 'Yes' if result.is_published else 'No',
                'Upload Date': result.uploaded_at.strftime('%Y-%m-%d %H:%M')
            })
        
        df = pd.DataFrame(data)
        
        # Create Excel file
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Results', index=False)
            
            # Add grading scale sheet
            grading_data = [
                {'Percentage': '90-100%', 'Grade': 'A+', 'Remarks': 'Outstanding', 'GPA': '4.0'},
                {'Percentage': '80-90%', 'Grade': 'A', 'Remarks': 'Excellent', 'GPA': '3.6'},
                {'Percentage': '70-80%', 'Grade': 'B+', 'Remarks': 'Very Good', 'GPA': '3.2'},
                {'Percentage': '60-70%', 'Grade': 'B', 'Remarks': 'Good', 'GPA': '2.8'},
                {'Percentage': '50-60%', 'Grade': 'C+', 'Remarks': 'Above Average', 'GPA': '2.4'},
                {'Percentage': '40-50%', 'Grade': 'C', 'Remarks': 'Average', 'GPA': '2.0'},
                {'Percentage': '20-40%', 'Grade': 'D', 'Remarks': 'Below Average', 'GPA': '1.6'},
                {'Percentage': '1-20%', 'Grade': 'E', 'Remarks': 'Insufficient', 'GPA': '0.8'},
            ]
            grading_df = pd.DataFrame(grading_data)
            grading_df.to_excel(writer, sheet_name='Grading Scale', index=False)
        
        output.seek(0)
        
        # Create response
        filename = f'results_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
def bulk_save_results(request):
    """Save multiple results at once"""
    try:
        data = json.loads(request.body)
        results_data = data.get('results', [])
        
        saved_count = 0
        for result_data in results_data:
            try:
                # Create or update result
                result, created = Result.objects.get_or_create(
                    roll_number=result_data['roll_number'],
                    student_class=result_data['student_class'],
                    exam_type=result_data['exam_type'],
                    academic_year=result_data['academic_year'],
                    defaults={
                        'student_name': result_data['student_name'],
                        'total': result_data['total'],
                        'gpa': result_data['gpa'],
                        'percentage': result_data['percentage'],
                        'total_subjects': 5,
                        'passed_subjects': result_data.get('passed_subjects', 5),
                        'failed_subjects': result_data.get('failed_subjects', 0),
                    }
                )
                
                if not created:
                    # Update existing result
                    result.student_name = result_data['student_name']
                    result.total = result_data['total']
                    result.gpa = result_data['gpa']
                    result.percentage = result_data['percentage']
                    result.save()
                
                saved_count += 1
                
            except Exception as e:
                continue
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully saved {saved_count} results'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
def delete_result(request):
    """Delete a single result by ID"""
    try:
        result_id = request.POST.get('result_id')
        if not result_id:
            return JsonResponse({'success': False, 'error': 'Result ID is required'})
        
        from .models import Result
        try:
            result = Result.objects.get(id=result_id)
            result.delete()
            return JsonResponse({'success': True, 'message': 'Result deleted successfully'})
        except Result.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Result not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# Enhanced Result Management API Views

class SubjectListCreate(generics.ListCreateAPIView):
    """API endpoint for listing and creating subjects"""
    queryset = Subject.objects.filter(is_active=True)
    serializer_class = SubjectSerializer
    permission_classes = [permissions.IsAdminUser]

class SubjectRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    """API endpoint for retrieving, updating, and deleting subjects"""
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer
    permission_classes = [permissions.IsAdminUser]

class MarksListCreate(generics.ListCreateAPIView):
    """API endpoint for listing and creating marks"""
    serializer_class = MarksSerializer
    permission_classes = [permissions.IsAdminUser]
    
    def get_queryset(self):
        queryset = Marks.objects.all()
        student_id = self.request.query_params.get('student_id', None)
        subject_id = self.request.query_params.get('subject_id', None)
        exam_type = self.request.query_params.get('exam_type', None)
        academic_year = self.request.query_params.get('academic_year', None)
        
        if student_id:
            queryset = queryset.filter(student_id=student_id)
        if subject_id:
            queryset = queryset.filter(subject_id=subject_id)
        if exam_type:
            queryset = queryset.filter(exam_type=exam_type)
        if academic_year:
            queryset = queryset.filter(academic_year=academic_year)
            
        return queryset

class MarksRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    """API endpoint for retrieving, updating, and deleting marks"""
    queryset = Marks.objects.all()
    serializer_class = MarksSerializer
    permission_classes = [permissions.IsAdminUser]

class StudentResultDetail(generics.RetrieveAPIView):
    """API endpoint for getting complete student result"""
    queryset = Student.objects.all()
    serializer_class = StudentResultSerializer
    permission_classes = [permissions.IsAdminUser]
    lookup_field = 'symbol_number'

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_subject_api(request):
    """API endpoint for adding a new subject"""
    try:
        name = request.data.get('name')
        credit_hour = request.data.get('credit_hour', 1.0)
        description = request.data.get('description', '')
        
        if not name:
            return Response({'error': 'Subject name is required'}, status=400)
        
        # Check if subject already exists (case-insensitive)
        existing_subjects = Subject.objects.filter(name__iexact=name)
        if existing_subjects.exists():
            # If multiple subjects found, use the first one
            existing_subject = existing_subjects.first()
            return Response({
                'error': f'Subject "{existing_subject.name}" already exists',
                'existing_subject': SubjectSerializer(existing_subject).data
            }, status=400)
        
        subject = Subject.objects.create(
            name=name,
            credit_hour=credit_hour,
            description=description
        )
        
        return Response({
            'success': True,
            'message': f'Subject "{name}" added successfully',
            'subject': SubjectSerializer(subject).data
        })
        
    except Exception as e:
        return Response({'error': str(e)}, status=500)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_student_api(request):
    """API endpoint for adding a new student"""
    try:
        full_name = request.data.get('full_name')
        student_class = request.data.get('student_class', '10')
        date_of_birth = request.data.get('date_of_birth')
        gender = request.data.get('gender', 'Other')
        email = request.data.get('email', '')
        phone_number = request.data.get('phone_number', '')
        address = request.data.get('address', '')
        parent_name = request.data.get('parent_name', '')
        parent_contact = request.data.get('parent_contact', '')
        
        if not full_name:
            return Response({'error': 'Student name is required'}, status=400)
        
        # Generate unique username
        base_username = full_name.lower().replace(' ', '')
        username = base_username
        counter = 1
        while Student.objects.filter(username=username).exists():
            username = f"{base_username}{counter}"
            counter += 1
        
        student = Student.objects.create(
            full_name=full_name,
            username=username,
            student_class=student_class,
            date_of_birth=date_of_birth,
            gender=gender,
            email=email,
            phone_number=phone_number,
            address=address,
            parent_name=parent_name,
            parent_contact=parent_contact
        )
        
        return Response({
            'success': True,
            'message': f'Student "{full_name}" added successfully',
            'student': {
                'id': student.id,
                'full_name': student.full_name,
                'symbol_number': student.symbol_number,
                'username': student.username
            }
        })
        
    except Exception as e:
        return Response({'error': str(e)}, status=500)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def submit_marks_api(request):
    """API endpoint for submitting marks for multiple students and subjects"""
    try:
        marks_data = request.data.get('marks', [])
        
        if not marks_data:
            return Response({'error': 'No marks data provided'}, status=400)
        
        created_count = 0
        updated_count = 0
        
        for mark_info in marks_data:
            student_symbol = mark_info.get('student_symbol')
            subject_name = mark_info.get('subject_name')
            exam_type = mark_info.get('exam_type', 'Final Term')
            academic_year = mark_info.get('academic_year', '2024-25')
            theory_marks = mark_info.get('theory_marks')
            practical_marks = mark_info.get('practical_marks')
            theory_total = mark_info.get('theory_total', 100)
            practical_total = mark_info.get('practical_total', 0)
            
            try:
                student = Student.objects.get(symbol_number=student_symbol)
                subject = Subject.objects.get(name=subject_name)
                
                # Create or update marks
                marks, created = Marks.objects.get_or_create(
                    student=student,
                    subject=subject,
                    exam_type=exam_type,
                    academic_year=academic_year,
                    defaults={
                        'theory_marks': theory_marks,
                        'practical_marks': practical_marks,
                        'theory_total': theory_total,
                        'practical_total': practical_total
                    }
                )
                
                if not created:
                    # Update existing marks
                    marks.theory_marks = theory_marks
                    marks.practical_marks = practical_marks
                    marks.theory_total = theory_total
                    marks.practical_total = practical_total
                    marks.save()
                    updated_count += 1
                else:
                    created_count += 1
                    
            except (Student.DoesNotExist, Subject.DoesNotExist) as e:
                continue
        
        return Response({
            'success': True,
            'message': f'Successfully processed {created_count + updated_count} marks (Created: {created_count}, Updated: {updated_count})'
        })
        
    except Exception as e:
        return Response({'error': str(e)}, status=500)

@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def student_result_api(request, symbol_number):
    """API endpoint for getting student result by symbol number"""
    try:
        student = Student.objects.get(symbol_number=symbol_number)
        serializer = StudentResultSerializer(student)
        return Response(serializer.data)
        
    except Student.DoesNotExist:
        return Response({'error': 'Student not found'}, status=404)
    except Exception as e:
        return Response({'error': str(e)}, status=500)

# New API endpoints for the enhanced result management system
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def subjects_search_api(request):
    # Check if user is staff/admin
    if not request.user.is_staff:
        return Response({'error': 'Access denied. Admin privileges required.'}, status=403)
    """API endpoint for searching subjects with live suggestions"""
    try:
        query = request.GET.get('query', '').strip()
        if not query:
            return Response({'subjects': []})
        
        # Case-insensitive search
        subjects = Subject.objects.filter(
            name__icontains=query,
            is_active=True
        ).order_by('name')[:10]  # Limit to 10 results
        
        serializer = SubjectSerializer(subjects, many=True)
        return Response({'subjects': serializer.data})
    except Exception as e:
        return Response({'error': str(e)}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def students_search_api(request):
    # Check if user is staff/admin
    if not request.user.is_staff:
        return Response({'error': 'Access denied. Admin privileges required.'}, status=403)
    """API endpoint for searching students by name or symbol number"""
    try:
        query = request.GET.get('query', '').strip()
        if not query:
            return Response({'students': []})
        
        # Search by name or symbol number
        students = Student.objects.filter(
            models.Q(full_name__icontains=query) | 
            models.Q(symbol_number__icontains=query)
        ).order_by('full_name')[:10]  # Limit to 10 results
        
        serializer = StudentSerializer(students, many=True)
        return Response({'students': serializer.data})
    except Exception as e:
        return Response({'error': str(e)}, status=500)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def marks_auto_save_api(request):
    # Check if user is staff/admin
    if not request.user.is_staff:
        return Response({'error': 'Access denied. Admin privileges required.'}, status=403)
    """API endpoint for auto-saving individual marks with debounce"""
    try:
        student_id = request.data.get('student_id')
        subject_id = request.data.get('subject_id')
        theory_marks = request.data.get('theory_marks')
        practical_marks = request.data.get('practical_marks')
        exam_type = request.data.get('exam_type', 'Final Term')
        academic_year = request.data.get('academic_year', '2024-25')
        
        if not all([student_id, subject_id]):
            return Response({'error': 'Student ID and Subject ID are required'}, status=400)
        
        # Get or create marks record
        marks, created = Marks.objects.get_or_create(
            student_id=student_id,
            subject_id=subject_id,
            exam_type=exam_type,
            academic_year=academic_year,
            defaults={
                'theory_marks': theory_marks,
                'practical_marks': practical_marks
            }
        )
        
        if not created:
            # Update existing record
            if theory_marks is not None:
                marks.theory_marks = theory_marks
            if practical_marks is not None:
                marks.practical_marks = practical_marks
            marks.save()
        
        serializer = MarksSerializer(marks)
        return Response({
            'success': True,
            'message': 'Marks saved successfully',
            'marks': serializer.data
        })
        
    except Exception as e:
        return Response({'error': str(e)}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def subject_marks_api(request, subject_id):
    """API endpoint for getting all marks for a specific subject"""
    try:
        exam_type = request.GET.get('exam_type', 'Final Term')
        academic_year = request.GET.get('academic_year', '2024-25')
        
        marks = Marks.objects.filter(
            subject_id=subject_id,
            exam_type=exam_type,
            academic_year=academic_year
        ).select_related('student', 'subject').order_by('student__full_name')
        
        serializer = MarksSerializer(marks, many=True)
        return Response({
            'marks': serializer.data,
            'subject': SubjectSerializer(marks.first().subject).data if marks.exists() else None
        })
        
    except Exception as e:
        return Response({'error': str(e)}, status=500)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generate_numeric_symbol_number_api(request):
    """API endpoint for generating unique numeric symbol numbers"""
    try:
        import random
        
        def generate_numeric_symbol():
            """Generate a unique 8-digit numeric symbol number"""
            while True:
                # Generate 8-digit number
                symbol = str(random.randint(10000000, 99999999))
                
                # Check if it already exists
                if not Student.objects.filter(symbol_number=symbol).exists():
                    return symbol
        
        symbol_number = generate_numeric_symbol()
        return Response({
            'success': True,
            'symbol_number': symbol_number
        })
        
    except Exception as e:
        return Response({'error': str(e)}, status=500)

def comprehensive_result_management_view(request):
    """View for the comprehensive result management system"""
    if not request.user.is_authenticated:
        return redirect('login_page')
    
    if not request.user.is_staff:
        return HttpResponseForbidden("Access denied. Admin privileges required.")
    
    return render(request, 'admin/core/result/comprehensive_result_management.html')

# ==================== EXPORT FUNCTIONALITY ====================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_comprehensive_results_pdf(request):
    """Export comprehensive results as PDF"""
    try:
        exam_type = request.GET.get('exam_type', 'Final Term')
        academic_year = request.GET.get('academic_year', '2024-25')
        subject_id = request.GET.get('subject_id')
        
        # Get marks data
        marks_queryset = Marks.objects.filter(
            exam_type=exam_type,
            academic_year=academic_year
        ).select_related('student', 'subject')
        
        if subject_id:
            marks_queryset = marks_queryset.filter(subject_id=subject_id)
        
        marks_queryset = marks_queryset.order_by('student__full_name')
        
        if not marks_queryset.exists():
            return Response({'error': 'No data found for the specified criteria'}, status=404)
        
        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Title
        title_text = f"Comprehensive Results Report - {exam_type} ({academic_year})"
        if subject_id:
            subject = Subject.objects.get(id=subject_id)
            title_text += f" - {subject.name}"
        
        elements.append(Paragraph(title_text, title_style))
        elements.append(Spacer(1, 20))
        
        # Table data
        table_data = [['Symbol No.', 'Student Name', 'Subject', 'Theory', 'Practical', 'Total', 'Grade', 'Status']]
        
        for mark in marks_queryset:
            total_marks = (mark.theory_marks or 0) + (mark.practical_marks or 0)
            grade = mark.grade or 'N/A'
            status = 'Pass' if mark.is_passed else 'Fail' if mark.is_passed is False else 'N/A'
            
            table_data.append([
                mark.student.symbol_number,
                mark.student.full_name,
                mark.subject.name,
                str(mark.theory_marks or 'N/A'),
                str(mark.practical_marks or 'N/A'),
                str(total_marks),
                grade,
                status
            ])
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        # Create response
        filename = f"comprehensive_results_{exam_type}_{academic_year}.pdf"
        if subject_id:
            subject = Subject.objects.get(id=subject_id)
            filename = f"results_{subject.name}_{exam_type}_{academic_year}.pdf"
        
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        return Response({'error': str(e)}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_comprehensive_results_excel(request):
    """Export comprehensive results as Excel"""
    try:
        exam_type = request.GET.get('exam_type', 'Final Term')
        academic_year = request.GET.get('academic_year', '2024-25')
        subject_id = request.GET.get('subject_id')
        
        # Get marks data
        marks_queryset = Marks.objects.filter(
            exam_type=exam_type,
            academic_year=academic_year
        ).select_related('student', 'subject')
        
        if subject_id:
            marks_queryset = marks_queryset.filter(subject_id=subject_id)
        
        marks_queryset = marks_queryset.order_by('student__full_name')
        
        if not marks_queryset.exists():
            return Response({'error': 'No data found for the specified criteria'}, status=404)
        
        # Create Excel workbook
        workbook = xlsxwriter.Workbook('temp_results.xlsx')
        worksheet = workbook.add_worksheet('Results')
        
        # Styles
        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'top',
            'fg_color': '#D7E4BC',
            'border': 1
        })
        
        cell_format = workbook.add_format({
            'text_wrap': True,
            'valign': 'top',
            'border': 1
        })
        
        # Headers
        headers = ['Symbol No.', 'Student Name', 'Subject', 'Theory', 'Practical', 'Total', 'Grade', 'Status']
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)
        
        # Data
        for row, mark in enumerate(marks_queryset, start=1):
            total_marks = (mark.theory_marks or 0) + (mark.practical_marks or 0)
            grade = mark.grade or 'N/A'
            status = 'Pass' if mark.is_passed else 'Fail' if mark.is_passed is False else 'N/A'
            
            worksheet.write(row, 0, mark.student.symbol_number, cell_format)
            worksheet.write(row, 1, mark.student.full_name, cell_format)
            worksheet.write(row, 2, mark.subject.name, cell_format)
            worksheet.write(row, 3, mark.theory_marks or 'N/A', cell_format)
            worksheet.write(row, 4, mark.practical_marks or 'N/A', cell_format)
            worksheet.write(row, 5, total_marks, cell_format)
            worksheet.write(row, 6, grade, cell_format)
            worksheet.write(row, 7, status, cell_format)
        
        # Auto-adjust column widths
        for col in range(len(headers)):
            worksheet.set_column(col, col, 15)
        
        workbook.close()
        
        # Read the file and create response
        with open('temp_results.xlsx', 'rb') as f:
            excel_data = f.read()
        
        # Clean up temporary file
        os.remove('temp_results.xlsx')
        
        # Create response
        filename = f"comprehensive_results_{exam_type}_{academic_year}.xlsx"
        if subject_id:
            subject = Subject.objects.get(id=subject_id)
            filename = f"results_{subject.name}_{exam_type}_{academic_year}.xlsx"
        
        response = HttpResponse(excel_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        return Response({'error': str(e)}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_student_result_card_pdf(request, symbol_number):
    """Export individual student result card as PDF"""
    try:
        student = Student.objects.get(symbol_number=symbol_number)
        exam_type = request.GET.get('exam_type', 'Final Term')
        academic_year = request.GET.get('academic_year', '2024-25')
        
        # Get all marks for the student
        marks = Marks.objects.filter(
            student=student,
            exam_type=exam_type,
            academic_year=academic_year
        ).select_related('subject').order_by('subject__name')
        
        if not marks.exists():
            return Response({'error': 'No results found for this student'}, status=404)
        
        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=20,
            alignment=TA_CENTER
        )
        
        # Title
        elements.append(Paragraph("NAWA PRATIVA SECONDARY SCHOOL", title_style))
        elements.append(Paragraph("STUDENT RESULT CARD", subtitle_style))
        elements.append(Spacer(1, 20))
        
        # Student Information
        student_info = [
            ['Student Name:', student.full_name],
            ['Symbol Number:', student.symbol_number],
            ['Class:', student.student_class],
            ['Exam Type:', exam_type],
            ['Academic Year:', academic_year],
            ['Date:', datetime.now().strftime('%Y-%m-%d')]
        ]
        
        student_table = Table(student_info, colWidths=[2*inch, 4*inch])
        student_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        elements.append(student_table)
        elements.append(Spacer(1, 20))
        
        # Results Table
        results_data = [['Subject', 'Theory', 'Practical', 'Total', 'Grade', 'Status']]
        
        total_marks = 0
        total_obtained = 0
        passed_subjects = 0
        
        for mark in marks:
            theory = mark.theory_marks or 0
            practical = mark.practical_marks or 0
            total = theory + practical
            grade = mark.grade or 'N/A'
            status = 'Pass' if mark.is_passed else 'Fail' if mark.is_passed is False else 'N/A'
            
            results_data.append([
                mark.subject.name,
                str(theory),
                str(practical),
                str(total),
                grade,
                status
            ])
            
            total_marks += (mark.subject.theory_total or 0) + (mark.subject.practical_total or 0)
            total_obtained += total
            if mark.is_passed:
                passed_subjects += 1
        
        # Calculate percentage
        percentage = (total_obtained / total_marks * 100) if total_marks > 0 else 0
        
        results_table = Table(results_data)
        results_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ]))
        
        elements.append(results_table)
        elements.append(Spacer(1, 20))
        
        # Summary
        summary_data = [
            ['Total Marks Obtained:', str(total_obtained)],
            ['Total Marks Possible:', str(total_marks)],
            ['Percentage:', f"{percentage:.2f}%"],
            ['Passed Subjects:', str(passed_subjects)],
            ['Total Subjects:', str(len(marks))],
            ['Overall Result:', 'PASS' if percentage >= 40 else 'FAIL']
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        elements.append(summary_table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        # Create response
        filename = f"result_card_{student.symbol_number}_{exam_type}_{academic_year}.pdf"
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Student.DoesNotExist:
        return Response({'error': 'Student not found'}, status=404)
    except Exception as e:
        return Response({'error': str(e)}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_student_result_card_excel(request, symbol_number):
    """Export individual student result card as Excel"""
    try:
        student = Student.objects.get(symbol_number=symbol_number)
        exam_type = request.GET.get('exam_type', 'Final Term')
        academic_year = request.GET.get('academic_year', '2024-25')
        
        # Get all marks for the student
        marks = Marks.objects.filter(
            student=student,
            exam_type=exam_type,
            academic_year=academic_year
        ).select_related('subject').order_by('subject__name')
        
        if not marks.exists():
            return Response({'error': 'No results found for this student'}, status=404)
        
        # Create Excel workbook
        workbook = xlsxwriter.Workbook('temp_result_card.xlsx')
        
        # Student info worksheet
        info_worksheet = workbook.add_worksheet('Student Info')
        
        # Styles
        title_format = workbook.add_format({
            'bold': True,
            'font_size': 16,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#D7E4BC',
            'border': 1
        })
        
        cell_format = workbook.add_format({
            'border': 1,
            'align': 'left'
        })
        
        # Title
        info_worksheet.merge_range('A1:B1', 'NAWA PRATIVA SECONDARY SCHOOL', title_format)
        info_worksheet.merge_range('A2:B2', 'STUDENT RESULT CARD', title_format)
        
        # Student Information
        student_info = [
            ['Student Name:', student.full_name],
            ['Symbol Number:', student.symbol_number],
            ['Class:', student.student_class],
            ['Exam Type:', exam_type],
            ['Academic Year:', academic_year],
            ['Date:', datetime.now().strftime('%Y-%m-%d')]
        ]
        
        for row, (label, value) in enumerate(student_info, start=4):
            info_worksheet.write(row, 0, label, header_format)
            info_worksheet.write(row, 1, value, cell_format)
        
        # Results worksheet
        results_worksheet = workbook.add_worksheet('Results')
        
        # Headers
        headers = ['Subject', 'Theory', 'Practical', 'Total', 'Grade', 'Status']
        for col, header in enumerate(headers):
            results_worksheet.write(0, col, header, header_format)
        
        # Data
        total_marks = 0
        total_obtained = 0
        passed_subjects = 0
        
        for row, mark in enumerate(marks, start=1):
            theory = mark.theory_marks or 0
            practical = mark.practical_marks or 0
            total = theory + practical
            grade = mark.grade or 'N/A'
            status = 'Pass' if mark.is_passed else 'Fail' if mark.is_passed is False else 'N/A'
            
            results_worksheet.write(row, 0, mark.subject.name, cell_format)
            results_worksheet.write(row, 1, theory, cell_format)
            results_worksheet.write(row, 2, practical, cell_format)
            results_worksheet.write(row, 3, total, cell_format)
            results_worksheet.write(row, 4, grade, cell_format)
            results_worksheet.write(row, 5, status, cell_format)
            
            total_marks += (mark.subject.theory_total or 0) + (mark.subject.practical_total or 0)
            total_obtained += total
            if mark.is_passed:
                passed_subjects += 1
        
        # Calculate percentage
        percentage = (total_obtained / total_marks * 100) if total_marks > 0 else 0
        
        # Summary
        summary_start_row = len(marks) + 3
        summary_data = [
            ['Total Marks Obtained:', total_obtained],
            ['Total Marks Possible:', total_marks],
            ['Percentage:', f"{percentage:.2f}%"],
            ['Passed Subjects:', passed_subjects],
            ['Total Subjects:', len(marks)],
            ['Overall Result:', 'PASS' if percentage >= 40 else 'FAIL']
        ]
        
        for row, (label, value) in enumerate(summary_data, start=summary_start_row):
            results_worksheet.write(row, 0, label, header_format)
            results_worksheet.write(row, 1, value, cell_format)
        
        # Auto-adjust column widths
        for worksheet in [info_worksheet, results_worksheet]:
            for col in range(10):
                worksheet.set_column(col, col, 15)
        
        workbook.close()
        
        # Read the file and create response
        with open('temp_result_card.xlsx', 'rb') as f:
            excel_data = f.read()
        
        # Clean up temporary file
        os.remove('temp_result_card.xlsx')
        
        # Create response
        filename = f"result_card_{student.symbol_number}_{exam_type}_{academic_year}.xlsx"
        response = HttpResponse(excel_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Student.DoesNotExist:
        return Response({'error': 'Student not found'}, status=404)
    except Exception as e:
        return Response({'error': str(e)}, status=500)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def grade_calculator_api(request):
    """Calculate grade based on percentage"""
    try:
        percentage = float(request.data.get('percentage', 0))
        
        if percentage < 0 or percentage > 100:
            return Response({'error': 'Percentage must be between 0 and 100'}, status=400)
        
        # Grade calculation logic
        if percentage >= 90:
            grade = 'A+'
            grade_point = 4.0
            status = 'Excellent'
        elif percentage >= 80:
            grade = 'A'
            grade_point = 3.6
            status = 'Very Good'
        elif percentage >= 70:
            grade = 'B+'
            grade_point = 3.2
            status = 'Good'
        elif percentage >= 60:
            grade = 'B'
            grade_point = 2.8
            status = 'Satisfactory'
        elif percentage >= 50:
            grade = 'C+'
            grade_point = 2.4
            status = 'Average'
        elif percentage >= 40:
            grade = 'C'
            grade_point = 2.0
            status = 'Pass'
        else:
            grade = 'F'
            grade_point = 0.0
            status = 'Fail'
        
        return Response({
            'percentage': percentage,
            'grade': grade,
            'grade_point': grade_point,
            'status': status,
            'passed': percentage >= 40
        })
        
    except ValueError:
        return Response({'error': 'Invalid percentage value'}, status=400)
    except Exception as e:
        return Response({'error': str(e)}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def print_result_card_view(request, symbol_number):
    """View for printing result card"""
    try:
        student = Student.objects.get(symbol_number=symbol_number)
        exam_type = request.GET.get('exam_type', 'Final Term')
        academic_year = request.GET.get('academic_year', '2024-25')
        
        # Get all marks for the student
        marks = Marks.objects.filter(
            student=student,
            exam_type=exam_type,
            academic_year=academic_year
        ).select_related('subject').order_by('subject__name')
        
        if not marks.exists():
            return Response({'error': 'No results found for this student'}, status=404)
        
        # Calculate totals
        total_marks = 0
        total_obtained = 0
        passed_subjects = 0
        
        for mark in marks:
            theory = mark.theory_marks or 0
            practical = mark.practical_marks or 0
            total = theory + practical
            
            total_marks += (mark.subject.theory_total or 0) + (mark.subject.practical_total or 0)
            total_obtained += total
            if mark.is_passed:
                passed_subjects += 1
        
        percentage = (total_obtained / total_marks * 100) if total_marks > 0 else 0
        
        context = {
            'student': student,
            'marks': marks,
            'exam_type': exam_type,
            'academic_year': academic_year,
            'total_obtained': total_obtained,
            'total_marks': total_marks,
            'percentage': percentage,
            'passed_subjects': passed_subjects,
            'total_subjects': len(marks),
            'overall_result': 'PASS' if percentage >= 40 else 'FAIL'
        }
        
        return render(request, 'admin/core/result/print_result_card.html', context)
        
    except Student.DoesNotExist:
        return Response({'error': 'Student not found'}, status=404)
    except Exception as e:
        return Response({'error': str(e)}, status=500)

# ==================== EMAIL NOTIFICATIONS & BULK OPERATIONS ====================

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def send_result_notification_email(request, symbol_number):
    """Send result notification email to parent/student"""
    try:
        student = Student.objects.get(symbol_number=symbol_number)
        
        # For GET requests, use default values for testing
        if request.method == 'GET':
            exam_type = 'Final Term'
            academic_year = '2024-25'
            recipient_email = student.email or student.parent_contact
            recipient_type = 'student'
            custom_subject = None
            custom_message = None
        else:
            # For POST requests, get data from request
            exam_type = request.data.get('exam_type', 'Final Term')
            academic_year = request.data.get('academic_year', '2024-25')
            recipient_email = request.data.get('recipient_email')
            recipient_type = request.data.get('recipient_type', 'parent')  # 'parent' or 'student'
            custom_subject = request.data.get('custom_subject')
            custom_message = request.data.get('custom_message')
        
        # Use student's email if no recipient email provided
        if not recipient_email:
            recipient_email = student.email or student.parent_contact
        
        if not recipient_email:
            return Response({
                'success': False,
                'error': 'No email address found for this student',
                'message': 'Please add an email address for the student first.'
            }, status=400)
        
        # Get results from the Result model (not Marks model)
        results = Result.objects.filter(
            student_name=student.full_name,
            exam_type=exam_type,
            academic_year=academic_year
        ).order_by('id')
        
        if not results.exists():
            return Response({
                'success': False,
                'error': 'No results found for this student',
                'message': f'No results found for {student.full_name} in {exam_type} ({academic_year})'
            }, status=404)
        
        # Get the first result (assuming one result per student per exam)
        result = results.first()
        
        # Prepare email context using Result model data
        context = {
            'student': student,
            'marks': [],  # We'll use result data instead of marks
            'exam_type': exam_type,
            'academic_year': academic_year,
            'total_obtained': result.total,
            'total_marks': result.total,  # Assuming total is the max possible
            'percentage': result.percentage,
            'gpa': result.gpa or 0.0,  # Include GPA from Result model
            'passed_subjects': result.passed_subjects,
            'total_subjects': result.total_subjects,
            'overall_result': 'PASS' if result.percentage >= 40 else 'FAIL',
            'parent_name': student.parent_name or f"Parent of {student.full_name}",
            'recipient_email': recipient_email,
            'generated_date': datetime.now().strftime('%Y-%m-%d %H:%M'),
            'result_url': f"{request.scheme}://{request.get_host()}/print-result-card/{student.symbol_number}/?exam_type={exam_type}&academic_year={academic_year}"
        }
        
        # Send email synchronously for better error handling
        try:
            # Render email templates
            html_content = render_to_string('emails/result_notification.html', context)
            text_content = render_to_string('emails/result_notification.txt', context)
            
            # Use custom subject if provided, otherwise use default
            if custom_subject:
                subject = custom_subject
            else:
                subject = f"Academic Excellence Report - {student.full_name} ({exam_type})"
            
            from_email = settings.DEFAULT_FROM_EMAIL
            
            email = EmailMultiAlternatives(
                subject=subject,
                body=text_content,
                from_email=from_email,
                to=[recipient_email]
            )
            email.attach_alternative(html_content, "text/html")
            
            # Send email
            email.send()
            
            # Log the successful email sending
            print(f"✅ SUCCESS: Result notification email sent to {recipient_email} for student {student.full_name}")
            
            return Response({
                'success': True,
                'message': f'✅ Academic Excellence Report sent successfully to {recipient_email}',
                'student_name': student.full_name,
                'exam_type': exam_type,
                'academic_year': academic_year,
                'recipient_email': recipient_email,
                'sent_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'result_data': {
                    'total': result.total,
                    'percentage': result.percentage,
                    'gpa': result.gpa,
                    'passed_subjects': result.passed_subjects,
                    'total_subjects': result.total_subjects
                }
            })
            
        except Exception as email_error:
            print(f"❌ ERROR: Failed to send email to {recipient_email}: {str(email_error)}")
            return Response({
                'success': False,
                'error': f'Failed to send email: {str(email_error)}',
                'message': 'Email sending failed. Please check your email configuration.',
                'student_name': student.full_name,
                'recipient_email': recipient_email
            }, status=500)
        
    except Student.DoesNotExist:
        return Response({
            'success': False,
            'error': 'Student not found',
            'message': f'Student with symbol number {symbol_number} not found.'
        }, status=404)
    except Exception as e:
        print(f"❌ ERROR: Unexpected error in send_result_notification_email: {str(e)}")
        return Response({
            'success': False,
            'error': str(e),
            'message': 'An unexpected error occurred while processing the request.'
        }, status=500)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bulk_send_result_notifications(request):
    """Bulk send result notifications to multiple students"""
    try:
        exam_type = request.data.get('exam_type', 'Final Term')
        academic_year = request.data.get('academic_year', '2024-25')
        student_ids = request.data.get('student_ids', [])
        recipient_type = request.data.get('recipient_type', 'parent')  # 'parent' or 'student'
        
        if not student_ids:
            return Response({'error': 'Student IDs are required'}, status=400)
        
        students = Student.objects.filter(id__in=student_ids)
        if not students.exists():
            return Response({'error': 'No valid students found'}, status=404)
        
        success_count = 0
        error_count = 0
        results = []
        
        for student in students:
            try:
                # Get student marks
                marks = Marks.objects.filter(
                    student=student,
                    exam_type=exam_type,
                    academic_year=academic_year
                ).select_related('subject')
                
                if not marks.exists():
                    results.append({
                        'student_name': student.full_name,
                        'symbol_number': student.symbol_number,
                        'status': 'error',
                        'message': 'No results found'
                    })
                    error_count += 1
                    continue
                
                # Calculate totals
                total_marks = 0
                total_obtained = 0
                
                for mark in marks:
                    theory = mark.theory_marks or 0
                    practical = mark.practical_marks or 0
                    total = theory + practical
                    
                    total_marks += (mark.subject.theory_total or 0) + (mark.subject.practical_total or 0)
                    total_obtained += total
                
                percentage = (total_obtained / total_marks * 100) if total_marks > 0 else 0
                
                # Determine recipient email
                recipient_email = None
                if recipient_type == 'parent' and student.parent_contact:
                    # Try to extract email from parent contact if it's an email
                    if '@' in student.parent_contact:
                        recipient_email = student.parent_contact
                elif recipient_type == 'student' and student.email:
                    recipient_email = student.email
                
                if not recipient_email:
                    results.append({
                        'student_name': student.full_name,
                        'symbol_number': student.symbol_number,
                        'status': 'error',
                        'message': f'No {recipient_type} email found'
                    })
                    error_count += 1
                    continue
                
                # Prepare email context
                context = {
                    'student': student,
                    'marks': marks,
                    'exam_type': exam_type,
                    'academic_year': academic_year,
                    'total_obtained': total_obtained,
                    'total_marks': total_marks,
                    'percentage': percentage,
                    'overall_result': 'PASS' if percentage >= 40 else 'FAIL',
                    'parent_name': student.parent_name or f"Parent of {student.full_name}",
                    'recipient_email': recipient_email,
                    'generated_date': datetime.now().strftime('%Y-%m-%d %H:%M'),
                    'result_url': f"{request.scheme}://{request.get_host()}/print-result-card/{student.symbol_number}/?exam_type={exam_type}&academic_year={academic_year}"
                }
                
                # Send email
                html_content = render_to_string('emails/result_notification.html', context)
                text_content = render_to_string('emails/result_notification.txt', context)
                
                subject = f"Result Notification - {student.full_name} ({exam_type})"
                from_email = settings.DEFAULT_FROM_EMAIL
                
                email = EmailMultiAlternatives(
                    subject=subject,
                    body=text_content,
                    from_email=from_email,
                    to=[recipient_email]
                )
                email.attach_alternative(html_content, "text/html")
                email.send()
                
                results.append({
                    'student_name': student.full_name,
                    'symbol_number': student.symbol_number,
                    'recipient_email': recipient_email,
                    'status': 'success',
                    'message': 'Email sent successfully'
                })
                success_count += 1
                
            except Exception as e:
                results.append({
                    'student_name': student.full_name,
                    'symbol_number': student.symbol_number,
                    'status': 'error',
                    'message': str(e)
                })
                error_count += 1
        
        return Response({
            'success': True,
            'message': f'Bulk email operation completed. {success_count} successful, {error_count} failed.',
            'summary': {
                'total_students': len(students),
                'success_count': success_count,
                'error_count': error_count
            },
            'results': results
        })
        
    except Exception as e:
        return Response({'error': str(e)}, status=500)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bulk_import_students(request):
    """Bulk import students from Excel file"""
    try:
        if 'file' not in request.FILES:
            return Response({'error': 'Excel file is required'}, status=400)
        
        excel_file = request.FILES['file']
        
        if not excel_file.name.endswith('.xlsx'):
            return Response({'error': 'Please upload an Excel (.xlsx) file'}, status=400)
        
        # Read Excel file
        import pandas as pd
        df = pd.read_excel(excel_file)
        
        required_columns = ['full_name', 'symbol_number', 'student_class', 'date_of_birth', 'gender']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            return Response({
                'error': f'Missing required columns: {", ".join(missing_columns)}'
            }, status=400)
        
        success_count = 0
        error_count = 0
        results = []
        
        for index, row in df.iterrows():
            try:
                # Check if student already exists
                if Student.objects.filter(symbol_number=row['symbol_number']).exists():
                    results.append({
                        'row': index + 2,  # Excel rows start from 1, and we have header
                        'symbol_number': row['symbol_number'],
                        'status': 'error',
                        'message': 'Student with this symbol number already exists'
                    })
                    error_count += 1
                    continue
                
                # Create student
                student_data = {
                    'full_name': row['full_name'],
                    'symbol_number': row['symbol_number'],
                    'student_class': row['student_class'],
                    'date_of_birth': pd.to_datetime(row['date_of_birth']).date(),
                    'gender': row['gender'],
                    'username': f"student_{row['symbol_number']}",
                }
                
                # Optional fields
                if 'email' in df.columns and pd.notna(row['email']):
                    student_data['email'] = row['email']
                if 'parent_name' in df.columns and pd.notna(row['parent_name']):
                    student_data['parent_name'] = row['parent_name']
                if 'parent_contact' in df.columns and pd.notna(row['parent_contact']):
                    student_data['parent_contact'] = row['parent_contact']
                if 'address' in df.columns and pd.notna(row['address']):
                    student_data['address'] = row['address']
                
                student = Student.objects.create(**student_data)
                
                results.append({
                    'row': index + 2,
                    'symbol_number': row['symbol_number'],
                    'student_name': row['full_name'],
                    'status': 'success',
                    'message': 'Student created successfully'
                })
                success_count += 1
                
            except Exception as e:
                results.append({
                    'row': index + 2,
                    'symbol_number': row.get('symbol_number', 'N/A'),
                    'status': 'error',
                    'message': str(e)
                })
                error_count += 1
        
        return Response({
            'success': True,
            'message': f'Bulk import completed. {success_count} students imported, {error_count} failed.',
            'summary': {
                'total_rows': len(df),
                'success_count': success_count,
                'error_count': error_count
            },
            'results': results
        })
        
    except Exception as e:
        return Response({'error': str(e)}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_student_template(request):
    """Download Excel template for student import"""
    try:
        # Create Excel workbook
        workbook = xlsxwriter.Workbook('temp_student_template.xlsx')
        worksheet = workbook.add_worksheet('Students')
        
        # Styles
        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'top',
            'fg_color': '#D7E4BC',
            'border': 1
        })
        
        # Headers
        headers = [
            'full_name*', 'symbol_number*', 'student_class*', 'date_of_birth*', 'gender*',
            'email', 'parent_name', 'parent_contact', 'address', 'phone_number'
        ]
        
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)
        
        # Sample data
        sample_data = [
            ['Ram Kumar Shrestha', '12345678', '10', '2005-01-15', 'Male', 
             'ram.shrestha@example.com', 'Hari Shrestha', 'hari.shrestha@example.com', 
             'Kathmandu, Nepal', '9876543210'],
            ['Sita Devi Thapa', '87654321', '10', '2005-06-20', 'Female',
             'sita.thapa@example.com', 'Ram Thapa', 'ram.thapa@example.com',
             'Lalitpur, Nepal', '9876543211']
        ]
        
        for row, data in enumerate(sample_data, start=1):
            for col, value in enumerate(data):
                worksheet.write(row, col, value)
        
        # Auto-adjust column widths
        for col in range(len(headers)):
            worksheet.set_column(col, col, 15)
        
        workbook.close()
        
        # Read the file and create response
        with open('temp_student_template.xlsx', 'rb') as f:
            excel_data = f.read()
        
        # Clean up temporary file
        os.remove('temp_student_template.xlsx')
        
        # Create response
        response = HttpResponse(excel_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="student_import_template.xlsx"'
        
        return response
        
    except Exception as e:
        return Response({'error': str(e)}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def analytics_dashboard_data(request):
    """Get analytics data for dashboard"""
    try:
        exam_type = request.GET.get('exam_type', 'Final Term')
        academic_year = request.GET.get('academic_year', '2024-25')
        
        # Get all marks for the specified exam
        marks = Marks.objects.filter(
            exam_type=exam_type,
            academic_year=academic_year
        ).select_related('student', 'subject')
        
        if not marks.exists():
            return Response({'error': 'No data found for the specified criteria'}, status=404)
        
        # Calculate analytics
        total_students = marks.values('student').distinct().count()
        total_subjects = marks.values('subject').distinct().count()
        
        # Performance statistics
        total_marks_possible = 0
        total_marks_obtained = 0
        pass_count = 0
        fail_count = 0
        
        for mark in marks:
            theory = mark.theory_marks or 0
            practical = mark.practical_marks or 0
            total = theory + practical
            
            total_marks_possible += (mark.subject.theory_total or 0) + (mark.subject.practical_total or 0)
            total_marks_obtained += total
            
            if mark.is_passed:
                pass_count += 1
            elif mark.is_passed is False:
                fail_count += 1
        
        overall_percentage = (total_marks_obtained / total_marks_possible * 100) if total_marks_possible > 0 else 0
        
        # Subject-wise performance
        subject_performance = []
        for subject in Subject.objects.filter(marks__exam_type=exam_type, marks__academic_year=academic_year).distinct():
            subject_marks = marks.filter(subject=subject)
            if subject_marks.exists():
                subject_total = sum((m.theory_marks or 0) + (m.practical_marks or 0) for m in subject_marks)
                subject_possible = sum((m.subject.theory_total or 0) + (m.subject.practical_total or 0) for m in subject_marks)
                subject_percentage = (subject_total / subject_possible * 100) if subject_possible > 0 else 0
                
                subject_performance.append({
                    'subject_name': subject.name,
                    'average_percentage': round(subject_percentage, 2),
                    'total_students': subject_marks.count()
                })
        
        # Grade distribution
        grade_distribution = {
            'A+': 0, 'A': 0, 'B+': 0, 'B': 0, 'C+': 0, 'C': 0, 'F': 0
        }
        
        for mark in marks:
            grade = mark.grade
            if grade in grade_distribution:
                grade_distribution[grade] += 1
        
        return Response({
            'success': True,
            'analytics': {
                'total_students': total_students,
                'total_subjects': total_subjects,
                'overall_percentage': round(overall_percentage, 2),
                'pass_count': pass_count,
                'fail_count': fail_count,
                'pass_rate': round((pass_count / (pass_count + fail_count) * 100), 2) if (pass_count + fail_count) > 0 else 0,
                'subject_performance': subject_performance,
                'grade_distribution': grade_distribution
            }
        })
        
    except Exception as e:
        return Response({'error': str(e)}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def email_students_list_api(request):
    """API endpoint to get all students for email notifications"""
    # Check if user is staff/admin
    if not request.user.is_staff:
        return Response({'error': 'Access denied. Admin privileges required.'}, status=403)
    
    try:
        # Get all students from database
        students = Student.objects.all().order_by('full_name')
        
        # Serialize students with only the fields needed for email
        students_data = []
        for student in students:
            students_data.append({
                'id': student.id,
                'name': student.full_name,
                'symbol_number': student.symbol_number,
                'email': getattr(student, 'email', ''),
                'student_class': getattr(student, 'student_class', ''),
            })
        
        return Response({
            'students': students_data,
            'total_count': len(students_data)
        })
        
    except Exception as e:
        return Response({'error': f'Error fetching students: {str(e)}'}, status=500)

# Enhanced Marks Entry System APIs
class StudentSearchView(APIView):
    """Search students by class and other criteria"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        class_name = request.query_params.get('class')
        search_term = request.query_params.get('search', '')
        
        queryset = Student.objects.all()
        
        if class_name:
            queryset = queryset.filter(student_class=class_name)
        
        if search_term:
            queryset = queryset.filter(
                models.Q(full_name__icontains=search_term) |
                models.Q(symbol_number__icontains=search_term) |
                models.Q(username__icontains=search_term)
            )
        
        serializer = StudentSerializer(queryset, many=True)
        return Response(serializer.data)

class SubjectMarksView(APIView):
    """Get marks for a specific subject and class"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        class_name = request.query_params.get('class')
        subject_name = request.query_params.get('subject')
        exam_type = request.query_params.get('exam_type')
        academic_year = request.query_params.get('academic_year')
        
        if not all([class_name, subject_name, exam_type, academic_year]):
            return Response({'error': 'Missing required parameters'}, status=400)
        
        try:
            subject = Subject.objects.get(name=subject_name)
            students = Student.objects.filter(student_class=class_name)
            
            marks_data = {}
            for student in students:
                try:
                    marks = Marks.objects.get(
                        student=student,
                        subject=subject,
                        exam_type=exam_type,
                        academic_year=academic_year
                    )
                    marks_data[student.symbol_number] = {
                        'theory_marks': marks.theory_marks,
                        'theory_total': marks.theory_total,
                        'practical_marks': marks.practical_marks,
                        'practical_total': marks.practical_total,
                        'total_marks': marks.total_marks,
                        'percentage': marks.percentage,
                        'grade': marks.grade,
                        'grade_point': marks.grade_point,
                        'is_passed': marks.is_passed
                    }
                except Marks.DoesNotExist:
                    marks_data[student.symbol_number] = {
                        'theory_marks': '',
                        'theory_total': 100,
                        'practical_marks': '',
                        'practical_total': 0,
                        'total_marks': '',
                        'percentage': '',
                        'grade': '',
                        'grade_point': '',
                        'is_passed': True
                    }
            
            return Response(marks_data)
            
        except Subject.DoesNotExist:
            return Response({'error': 'Subject not found'}, status=404)

class EnhancedResultView(APIView):
    """Enhanced result view with better filtering and search"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        class_name = request.query_params.get('class')
        exam_type = request.query_params.get('exam_type')
        academic_year = request.query_params.get('academic_year')
        search_term = request.query_params.get('search', '')
        
        queryset = Result.objects.all()
        
        if class_name:
            queryset = queryset.filter(student_class=class_name)
        if exam_type:
            queryset = queryset.filter(exam_type=exam_type)
        if academic_year:
            queryset = queryset.filter(academic_year=academic_year)
        if search_term:
            queryset = queryset.filter(
                models.Q(student_name__icontains=search_term) |
                models.Q(roll_number__icontains=search_term)
            )
        
        # Include related subject results
        queryset = queryset.prefetch_related('subjects')
        
        serializer = ResultSerializer(queryset, many=True)
        return Response(serializer.data)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def enhanced_bulk_save_marks(request):
    """Enhanced bulk save marks with auto-calculation and validation"""
    try:
        data = request.data
        class_name = data.get('class_name')
        exam_type = data.get('exam_type')
        academic_year = data.get('academic_year')
        subject = data.get('subject')
        marks_data = data.get('marks_data', {})
        
        if not all([class_name, exam_type, academic_year, subject, marks_data]):
            return Response({'error': 'Missing required fields'}, status=400)
        
        # Get or create subject
        subject_obj, created = Subject.objects.get_or_create(
            name=subject,
            defaults={'code': subject[:3].upper(), 'credit_hour': 4.0}
        )
        
        saved_count = 0
        errors = []
        
        # Process each student's marks
        for symbol_number, marks in marks_data.items():
            try:
                student = Student.objects.get(symbol_number=symbol_number, student_class=class_name)
                
                # Validate marks
                theory_marks = float(marks.get('theory_marks', 0)) if marks.get('theory_marks') else None
                theory_total = float(marks.get('theory_total', 100))
                practical_marks = float(marks.get('practical_marks', 0)) if marks.get('practical_marks') else None
                practical_total = float(marks.get('practical_total', 0))
                
                # Validate theory marks
                if theory_marks is not None and (theory_marks < 0 or theory_marks > theory_total):
                    errors.append(f"Invalid theory marks for {student.full_name}: {theory_marks}")
                    continue
                
                # Validate practical marks
                if practical_marks is not None and (practical_marks < 0 or practical_marks > practical_total):
                    errors.append(f"Invalid practical marks for {student.full_name}: {practical_marks}")
                    continue
                
                # Get or create marks record
                marks_obj, created = Marks.objects.get_or_create(
                    student=student,
                    subject=subject_obj,
                    exam_type=exam_type,
                    academic_year=academic_year,
                    defaults={
                        'theory_marks': theory_marks,
                        'theory_total': theory_total,
                        'practical_marks': practical_marks,
                        'practical_total': practical_total,
                    }
                )
                
                if not created:
                    # Update existing record
                    marks_obj.theory_marks = theory_marks
                    marks_obj.theory_total = theory_total
                    marks_obj.practical_marks = practical_marks
                    marks_obj.practical_total = practical_total
                
                # Save to trigger auto-calculation
                marks_obj.save()
                saved_count += 1
                
            except Student.DoesNotExist:
                errors.append(f"Student not found: {symbol_number}")
            except ValueError as e:
                errors.append(f"Invalid marks format for {symbol_number}: {str(e)}")
            except Exception as e:
                errors.append(f"Error saving marks for {symbol_number}: {str(e)}")
        
        return Response({
            'success': True,
            'message': f'Successfully saved marks for {saved_count} students',
            'saved_count': saved_count,
            'errors': errors
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_subjects_for_class(request):
    """Get available subjects for a specific class"""
    try:
        class_name = request.query_params.get('class')
        
        if not class_name:
            return Response({'error': 'Class parameter is required'}, status=400)
        
        # Get subjects that have been used for this class
        used_subjects = Marks.objects.filter(
            student__student_class=class_name
        ).values_list('subject__name', 'subject__code', 'subject__credit_hour').distinct()
        
        subjects = []
        for name, code, credit_hour in used_subjects:
            subjects.append({
                'name': name,
                'code': code,
                'credit_hour': credit_hour
            })
        
        # If no subjects found, return default subjects
        if not subjects:
            default_subjects = [
                {'name': 'English', 'code': 'ENG', 'credit_hour': 4.0},
                {'name': 'Nepali', 'code': 'NEP', 'credit_hour': 5.0},
                {'name': 'Mathematics', 'code': 'MATH', 'credit_hour': 5.0},
                {'name': 'Science', 'code': 'SCI', 'credit_hour': 5.0},
                {'name': 'Social Studies', 'code': 'SOC', 'credit_hour': 4.0},
                {'name': 'Computer', 'code': 'COMP', 'credit_hour': 3.0}
            ]
            subjects = default_subjects
        
        return Response({
            'success': True,
            'subjects': subjects
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=500)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def set_subjects_for_class(request):
    """Define subjects for a class by creating placeholder marks rows for all students in that class.
    Body: { class_name, exam_type, academic_year, subjects: [subject_name or {name, credit_hour}] }
    """
    try:
        data = request.data
        class_name = data.get('class_name')
        exam_type = data.get('exam_type', 'Final Term')
        academic_year = data.get('academic_year', '2024-25')
        subjects_input = data.get('subjects', [])

        if not class_name or not subjects_input:
            return Response({'error': 'class_name and subjects are required'}, status=400)

        # Normalize subject names
        subject_names = []
        for s in subjects_input:
            if isinstance(s, dict):
                subject_names.append(s.get('name'))
            else:
                subject_names.append(str(s))
        subject_names = [name for name in subject_names if name]
        if not subject_names:
            return Response({'error': 'No valid subjects provided'}, status=400)

        # Ensure Subject records exist
        from .models import Subject, Student, Marks
        subjects = []
        for name in subject_names:
            subj, _ = Subject.objects.get_or_create(
                name=name,
                defaults={'credit_hour': 4.0}
            )
            subjects.append(subj)

        students = Student.objects.filter(student_class=class_name)
        if not students.exists():
            return Response({'error': f'No students found for class {class_name}'}, status=404)

        created_or_existing = 0
        for student in students:
            for subj in subjects:
                _, created = Marks.objects.get_or_create(
                    student=student,
                    subject=subj,
                    exam_type=exam_type,
                    academic_year=academic_year,
                    defaults={'theory_total': 100, 'practical_total': 0}
                )
                created_or_existing += 1

        return Response({'success': True, 'records': created_or_existing})
    except Exception as e:
        return Response({'success': False, 'error': str(e)}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_marks_entry_progress(request):
    """Get progress of marks entry for a class"""
    try:
        class_name = request.query_params.get('class')
        exam_type = request.query_params.get('exam_type')
        academic_year = request.query_params.get('academic_year')
        
        if not all([class_name, exam_type, academic_year]):
            return Response({'error': 'Missing required parameters'}, status=400)
        
        # Get all students in the class
        students = Student.objects.filter(student_class=class_name)
        total_students = students.count()
        
        if total_students == 0:
            return Response({
                'success': True,
                'progress': {
                    'total_students': 0,
                    'subjects_with_marks': 0,
                    'total_subjects': 0,
                    'completion_percentage': 0
                }
            })
        
        # Get subjects that have marks for this class/exam/academic year
        subjects_with_marks = Marks.objects.filter(
            student__student_class=class_name,
            exam_type=exam_type,
            academic_year=academic_year
        ).values_list('subject__name', flat=True).distinct()
        
        # Get all subjects that should have marks (based on existing marks)
        all_subjects = Marks.objects.filter(
            student__student_class=class_name
        ).values_list('subject__name', flat=True).distinct()
        
        total_subjects = len(all_subjects) if all_subjects else 6  # Default to 6 subjects
        subjects_with_marks_count = len(subjects_with_marks)
        
        completion_percentage = (subjects_with_marks_count / total_subjects) * 100 if total_subjects > 0 else 0
        
        return Response({
            'success': True,
            'progress': {
                'total_students': total_students,
                'subjects_with_marks': subjects_with_marks_count,
                'total_subjects': total_subjects,
                'completion_percentage': round(completion_percentage, 2),
                'subjects_completed': list(subjects_with_marks),
                'subjects_pending': list(set(all_subjects) - set(subjects_with_marks))
            }
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=500)
